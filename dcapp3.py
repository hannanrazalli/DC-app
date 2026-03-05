import customtkinter as ctk
import psycopg2
import os
import threading
import shutil
import time
from tkinter import messagebox
from datetime import date, datetime, timedelta
from tkcalendar import DateEntry
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from PIL import Image, ImageTk 
from fpdf import FPDF

# --- APPEARANCE SETUP ---
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("dark-blue")

# --- CONFIGURATION (Kombinasi Kod 1 & Kod 2) ---
BASE_PATH = r"C:\Users\HP\Documents\[01] Document Control"
DB_CONFIG = {"dbname": "dcde", "user": "postgres", "password": "1234", "host": "localhost"}
MASTER_FILE = "Document Control.xlsx"
LOGO_FILENAME = "LMG Locomotive Logo.jpeg" 

# CONFIGURATION DARI KOD 2
SERVER_PATH = r"Y:\[04] ENGINEERING TEAM\[98] DOCUMENT CONTROL"
FILE_MAPPING = {
    "Document Control.xlsx": "All Projects Drawings Data.xlsx",
    "H10 Berapit.xlsx": "H10 BeraPit Drawing List.xlsx",
    "H10 TRC.xlsx": "H10 TRC Drawing List.xlsx",
    "M10.xlsx": "M10(N) Drawing List.xlsx",
    "N10.xlsx": "N10(N) Drawing List.xlsx",
    "Wheel Press Machine.xlsx": "Wheel Press Machine Drawing List.xlsx",
    "G10.xlsx": "G10 Drawing List.xlsx"
}

# --- PDF DRAWING CONFIGURATION ---
BASE_DRAWINGS_PATH = r"Y:\[04] ENGINEERING TEAM\[98] DOCUMENT CONTROL\[00] 2D Drawings - Signed Copy"

# --- TRANSMITTAL CONFIGURATION (FUNGSI BARU) ---
TRANSMITTAL_FILE_PATH = r"Y:\[04] ENGINEERING TEAM\[98] DOCUMENT CONTROL\[05] Transmittal Form DC\DC Transmittal List.xlsx"

# Transmittal Server Configurations (Dari Kod 2)
SERVER_BASE = r"\\192.168.2.3\Drawings"
DC_ROOT = os.path.join(SERVER_BASE, "[04] ENGINEERING TEAM", "[98] DOCUMENT CONTROL", "[05] Transmittal Form DC")
DC_LIST_FILE = os.path.join(DC_ROOT, "DC Transmittal List.xlsx")

# PROJECT CONFIGURATION
PROJECTS = ["H10 TRC", "H10 BeraPit", "M10", "N10", "G10", "Wheel Press Machine"]

PROJ_MAP = {
    "H10 TRC": "H10", 
    "H10 BeraPit": "H10", 
    "M10": "M10", 
    "N10": "N10", 
    "G10": "G10 Drawing List", 
    "Wheel Press Machine": "WPM"
}

REVERSE_PROJ_MAP = {}
for k, v in PROJ_MAP.items():
    if v not in REVERSE_PROJ_MAP: 
        REVERSE_PROJ_MAP[v] = k

ASSEMBLIES = ["Bogie", "Underframe", "Cabin", "Engine Hood", "Radiator Hood", "Muffler", "Gear Case", "Water Tank", "Battery Box", "Fuel Tank", "Sandbox"]
ENGINEER_LIST = ["Baskaran", "Sathish", "Harrison", "Hannan", "Gokul", "Vimal", "Ram", "Vishwa", "Bruno", "Satyanarayana"]
REMARKS_LIST = ["New", "Revised", "Re-Release", "-"]
# BATCH_LIST dari kod anda
BATCH_LIST = ["-", "1", "2", "N", "R"]
MASTER_HEADERS = ["Project", "Country", "Batch", "Main Assembly", "Drawing Name", "Part Number", "Revision", "Total Sheets", "Engineer", "Date Approved", "Remarks"]

# User Credentials
USERS = {
    "Design Engineering": {"pin": "1112", "role": "ENG"},
    "Document Control": {"pin": "1234", "role": "DC"}
}

# --- COLORS & FONTS ---
COLOR_PRIMARY = "#2C3E50"    
COLOR_ACCENT = "#3498DB"     
COLOR_SUCCESS = "#27AE60"    
COLOR_DANGER = "#C0392B"     
COLOR_WARNING = "#F39C12"    
COLOR_INFO = "#1ABC9C"       
COLOR_BG = "#ECF0F1"         
COLOR_CARD = "#FFFFFF"       
COLOR_TEXT = "#34495E"       

FONT_HEADER = ("Segoe UI", 24, "bold")
FONT_SECTION = ("Segoe UI", 16, "bold")
FONT_LABEL = ("Segoe UI", 12, "bold")
FONT_INPUT = ("Segoe UI", 13)
FONT_FEEDBACK = ("Segoe UI", 16, "bold") 
FONT_ERROR = ("Segoe UI", 10, "bold")

CREAM_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def try_int(val):
    try: return int(val)
    except: return -1

# --- PDF GENERATOR ---
class TransmittalPDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "DOCUMENT TRANSMITTAL NOTE", 0, 1, "C")
        self.ln(5)

    def generate_table(self, header, data, proj_info):
        # Landscape Mode width ~ 277mm
        widths = [10, 40, 80, 15, 15, 25, 92]
        
        self.set_font("Arial", "B", 12)
        self.set_fill_color(144, 238, 144) 
        self.cell(0, 10, f"PROJECT: {proj_info['proj']}   |   MAIN ASSY: {proj_info['assy']}", 1, 1, 'C', 1)
        self.ln(2)

        self.set_font("Arial", "B", 8)
        self.set_fill_color(220, 220, 220)
        for i, h in enumerate(header):
            self.cell(widths[i], 8, h, 1, 0, 'C', 1)
        self.ln()
        
        self.set_font("Arial", "", 8)
        for row in data:
            if self.get_y() > 180: self.add_page() # Adjust for landscape
            for i, item in enumerate(row):
                align = 'L' if i in [2, 6] else 'C' # Left align drawing name and remarks
                self.cell(widths[i], 7, str(item), 1, 0, align)
            self.ln()

class IntegratedApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("Light")
        self.title("LMG Engineering - Integrated Data Entry & Transmittal System")
        # UBAHSUAI 1: TINGGIKAN VERTICAL UI KE 1050
        self.geometry("1250x1050") 
        self.minsize(1000, 950) 
        self.configure(fg_color=COLOR_BG)
        
        try:
            if os.path.exists(LOGO_FILENAME):
                self.window_icon = ImageTk.PhotoImage(file=LOGO_FILENAME)
                self.wm_iconphoto(False, self.window_icon)
        except Exception as e:
            print(f"Warning: Could not set window icon. {e}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) 
        
        self.current_user = None
        self.current_role = None
        self.project_scrolls = {}

        self.withdraw()
        self.show_login()

    def show_login(self):
        self.login_win = ctk.CTkToplevel(self)
        self.login_win.title("Security Login")
        self.login_win.geometry("400x500")
        self.login_win.attributes("-topmost", True)
        self.login_win.protocol("WM_DELETE_WINDOW", self.quit)

        ctk.CTkLabel(self.login_win, text="SYSTEM LOGIN", font=("Segoe UI", 28, "bold"), text_color="#2C3E50").pack(pady=40)
        
        self.user_var = ctk.StringVar(value="Design Engineering")
        ctk.CTkOptionMenu(self.login_win, values=list(USERS.keys()), variable=self.user_var, height=40, font=("Segoe UI", 14)).pack(pady=10)
        
        self.pin_ent = ctk.CTkEntry(self.login_win, placeholder_text="Enter PIN Code", show="*", height=50, font=("Consolas", 24), justify="center")
        self.pin_ent.pack(pady=20, padx=60)
        self.pin_ent.focus_set()
        
        self.pin_ent.bind("<Return>", lambda e: self.verify_login())
        ctk.CTkButton(self.login_win, text="LOGIN", command=self.verify_login, height=50, font=("Segoe UI", 16, "bold"), fg_color="#2C3E50").pack(pady=30)

    def verify_login(self):
        u = self.user_var.get()
        p = self.pin_ent.get()
        if USERS.get(u)["pin"] == p:
            self.current_user = u
            self.current_role = USERS[u]["role"]
            self.login_win.destroy()
            self.deiconify()
            self.setup_ui()
        else:
            messagebox.showerror("Error", "Wrong PIN! Please try again.")

    def setup_ui(self):
        self.header_frame = ctk.CTkFrame(self, fg_color=COLOR_PRIMARY, corner_radius=0, height=85)
        self.header_frame.grid(row=0, column=0, sticky="ew")
        self.header_frame.grid_propagate(False)

        self.title_label = ctk.CTkLabel(self.header_frame, text="ENGINEERING DOCUMENT CONTROL", font=FONT_HEADER, text_color="white")
        self.title_label.pack(side="left", padx=25, pady=15)
        
        self.subtitle_label = ctk.CTkLabel(self.header_frame, text="|  Integrated System", font=("Segoe UI", 14), text_color="#BDC3C7")
        self.subtitle_label.pack(side="left", pady=20)

        self.right_side_container = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        self.right_side_container.pack(side="right", padx=30)

        if self.current_role == "ENG":
            self.publish_ui_frame = ctk.CTkFrame(self.right_side_container, fg_color="transparent")
            self.publish_ui_frame.pack(side="left", padx=(0, 20))

            self.btn_publish = ctk.CTkButton(self.publish_ui_frame, text="PUBLISH TO SERVER", 
                                           fg_color=COLOR_INFO, hover_color="#16A085", 
                                           font=("Segoe UI", 12, "bold"), height=32,
                                           command=self.publish_to_server_thread)
            self.btn_publish.pack(pady=(0, 2))

            self.lbl_last_sync = ctk.CTkLabel(self.publish_ui_frame, text="Last Sync: Never", font=("Segoe UI", 10), text_color="#BDC3C7")
            self.lbl_last_sync.pack()

        ctk.CTkButton(self.right_side_container, text="LOGOUT", width=90, fg_color="#C0392B", font=("Segoe UI", 13, "bold"), command=self.logout).pack(side="left", padx=20)

        try:
            if os.path.exists(LOGO_FILENAME):
                pil_img = Image.open(LOGO_FILENAME)
                target_height = 40
                aspect_ratio = pil_img.width / pil_img.height
                target_width = int(target_height * aspect_ratio)
                self.header_logo = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(target_width, target_height))
                self.logo_label = ctk.CTkLabel(self.right_side_container, text="", image=self.header_logo)
                self.logo_label.pack(side="right")
        except Exception as e:
            print(f"Error loading logo: {e}")

        self.tabview = ctk.CTkTabview(self, fg_color=COLOR_CARD, border_width=1, border_color="#BDC3C7", corner_radius=15)
        self.tabview.grid(row=1, column=0, sticky="nsew", padx=20, pady=(10, 20))
        
        if self.current_role == "ENG":
            self.tab_entry = self.tabview.add("DATA ENTRY")
            self.tab_drawings = self.tabview.add("VIEW DRAWINGS")
            self.tab_trans = self.tabview.add("DRAWING TRANSMITTAL")
            
            self.setup_data_entry_tab()
            self.setup_view_drawings_tab()
            self.setup_hannan_dashboard()
        else:
            self.tab_drawings = self.tabview.add("VIEW DRAWINGS")
            self.tab_trans = self.tabview.add("DRAWING TRANSMITTAL")
            
            self.setup_view_drawings_tab()
            self.setup_tasya_dashboard()
            self.tabview.set("DRAWING TRANSMITTAL") # DEFAULT TAB DC

        self.status_bar = ctk.CTkFrame(self, height=30, fg_color="#BDC3C7", corner_radius=0)
        self.status_bar.grid(row=2, column=0, sticky="ew")
        self.status_label = ctk.CTkLabel(self.status_bar, text="System Ready", font=("Consolas", 11), text_color="#2C3E50")
        self.status_label.pack(side="left", padx=20)

    def logout(self):
        if messagebox.askyesno("Logout", "Do you want to exit?"):
            self.current_user = None
            for w in self.winfo_children(): w.destroy()
            self.withdraw()
            self.show_login()

    # =========================================================================
    # TAB 1: DATA ENTRY
    # =========================================================================
    def setup_data_entry_tab(self):
        self.tab_entry.columnconfigure((0, 1), weight=1)
        
        self.create_section_header("1. PROJECT DETAILS", row=0, parent=self.tab_entry)
        self.add_input_field(label="Project Name", row=1, col=0, parent=self.tab_entry)
        self.proj_v = ctk.StringVar(value=PROJECTS[0])
        self.proj_drop = ctk.CTkOptionMenu(self.tab_entry, values=PROJECTS, variable=self.proj_v, command=self.update_logic, 
                                           fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.proj_drop.grid(row=2, column=0, padx=20, pady=(5, 15), sticky="ew")

        self.add_input_field(label="Batch Code", row=1, col=1, parent=self.tab_entry)
        self.batch_v = ctk.StringVar(value="-")
        self.batch_drop = ctk.CTkOptionMenu(self.tab_entry, values=BATCH_LIST, variable=self.batch_v,
                                            fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.batch_drop.grid(row=2, column=1, padx=20, pady=(5, 15), sticky="ew")

        self.create_section_header("2. DRAWING & TECHNICAL INFORMATION", row=3, parent=self.tab_entry)
        self.add_input_field(label="Main Assembly", row=4, col=0, parent=self.tab_entry)
        self.assembly_v = ctk.StringVar(value=ASSEMBLIES[0])
        self.assembly_drop = ctk.CTkOptionMenu(self.tab_entry, values=ASSEMBLIES, variable=self.assembly_v, 
                                                fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.assembly_drop.grid(row=5, column=0, padx=20, pady=(5, 10), sticky="ew")

        self.add_input_field(label="Drawing Name", row=4, col=1, parent=self.tab_entry)
        self.draw_ent = ctk.CTkEntry(self.tab_entry, font=FONT_INPUT, height=32, placeholder_text="Full drawing title")
        self.draw_ent.grid(row=5, column=1, padx=20, pady=(5, 0), sticky="ew")
        self.draw_ent.bind("<KeyRelease>", lambda e: [self.to_uppercase(e, self.draw_ent), self.clear_field_error("draw")])
        self.err_draw = ctk.CTkLabel(self.tab_entry, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_draw.grid(row=6, column=1, padx=25, sticky="w")

        self.add_input_field(label="Part Number", row=7, col=0, parent=self.tab_entry)
        self.part_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.part_frame.grid(row=8, column=0, padx=20, pady=(5, 0), sticky="ew")
        self.part_ent = ctk.CTkEntry(self.part_frame, font=FONT_INPUT, height=32, placeholder_text="e.g. H10-100-001")
        self.part_ent.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.part_ent.bind("<KeyRelease>", lambda e: [self.on_part_input(e), self.clear_field_error("part")])
        
        self.part_ent.bind("<Return>", lambda e: self.check_part_existence_thread())
        
        self.btn_check_part = ctk.CTkButton(self.part_frame, text="Search", width=60, height=32, 
                                            fg_color="#7f8c8d", hover_color="#95a5a6", 
                                            command=self.check_part_existence_thread)
        self.btn_check_part.pack(side="right")
        self.err_part = ctk.CTkLabel(self.tab_entry, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_part.grid(row=9, column=0, padx=25, sticky="w")

        self.sub_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.sub_frame.grid(row=7, column=1, rowspan=3, padx=20, pady=0, sticky="ew")
        self.sub_frame.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkLabel(self.sub_frame, text="Revision", font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=0, column=0, sticky="w")
        self.rev_ent = ctk.CTkEntry(self.sub_frame, font=FONT_INPUT, height=32)
        self.rev_ent.grid(row=1, column=0, padx=(0, 10), pady=(5, 0), sticky="ew")
        self.rev_ent.bind("<KeyRelease>", lambda e: [self.auto_remark_logic(e), self.clear_field_error("rev")])
        self.err_rev = ctk.CTkLabel(self.sub_frame, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_rev.grid(row=2, column=0, padx=(0, 10), sticky="w")

        ctk.CTkLabel(self.sub_frame, text="Total Sheets", font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=0, column=1, sticky="w")
        self.total_ent = ctk.CTkEntry(self.sub_frame, font=FONT_INPUT, height=32)
        self.total_ent.grid(row=1, column=1, padx=(10, 0), pady=(5, 0), sticky="ew")
        self.total_ent.bind("<KeyRelease>", lambda e: self.clear_field_error("total"))
        self.err_total = ctk.CTkLabel(self.sub_frame, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_total.grid(row=2, column=1, padx=(10, 0), sticky="w")

        self.create_section_header("3. APPROVAL & STATUS", row=10, parent=self.tab_entry)
        self.add_input_field(label="Responsible Engineer", row=11, col=0, parent=self.tab_entry)
        self.eng_v = ctk.StringVar(value=ENGINEER_LIST[0])
        # Pengekalan saiz sticky ew mengikut permintaan
        self.eng_drop = ctk.CTkOptionMenu(self.tab_entry, values=ENGINEER_LIST, variable=self.eng_v, 
                                          fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.eng_drop.grid(row=12, column=0, padx=20, pady=(5, 10), sticky="ew")

        self.add_input_field(label="Date Approved", row=11, col=1, parent=self.tab_entry)
        self.date_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.date_frame.grid(row=12, column=1, padx=20, pady=(5, 10), sticky="w")
        self.date_picker = DateEntry(self.date_frame, width=20, background=COLOR_PRIMARY, date_pattern='yyyy-mm-dd')
        self.date_picker.pack(side="left", padx=(0, 15), ipady=3)
        ctk.CTkButton(self.date_frame, text="Set Today", width=100, height=30, fg_color="#95a5a6", command=self.set_today).pack(side="left")

        self.add_input_field(label="Remarks", row=13, col=0, parent=self.tab_entry)
        self.remark_v = ctk.StringVar(value="New")
        # Pengekalan saiz sticky ew mengikut permintaan
        self.remark_drop = ctk.CTkOptionMenu(self.tab_entry, values=REMARKS_LIST, variable=self.remark_v, 
                                             fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.remark_drop.grid(row=14, column=0, padx=20, pady=(5, 15), sticky="ew")

        # UBAHSUAI: REASON FOR REVISION TEXTBOX
        self.add_input_field(label="Reason for Revision (Compulsory if Revised):", row=13, col=1, parent=self.tab_entry)
        self.txt_reason = ctk.CTkTextbox(self.tab_entry, height=80, font=FONT_INPUT)
        self.txt_reason.grid(row=14, column=1, rowspan=2, padx=20, pady=(5, 15), sticky="ew")

        self.feedback_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent", height=50)
        self.feedback_frame.grid(row=16, column=0, columnspan=2, pady=(10, 5), padx=20, sticky="ew")
        self.lbl_feedback = ctk.CTkLabel(self.feedback_frame, text="", font=FONT_FEEDBACK, text_color="white", corner_radius=6)
        self.lbl_feedback.pack(fill="both", ipady=10)

        self.btn_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.btn_frame.grid(row=17, column=0, columnspan=2, pady=(10, 20), padx=20, sticky="ew")
        self.btn_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        ctk.CTkButton(self.btn_frame, text="SUBMIT DATA", fg_color=COLOR_SUCCESS, hover_color="#219150", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.submit).grid(row=0, column=0, padx=(0, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="CLEAR FORM", fg_color=COLOR_DANGER, hover_color="#A93226", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.clear_all).grid(row=0, column=1, padx=(5, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="OPEN FOLDER", fg_color=COLOR_WARNING, hover_color="#D35400", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.open_folder).grid(row=0, column=2, padx=(5, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="OPEN EXCEL", fg_color=COLOR_INFO, hover_color="#16A085", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.open_project_file).grid(row=0, column=3, padx=(5, 0), sticky="ew")


    # =========================================================================
    # TAB 2: VIEW DRAWINGS
    # =========================================================================
    def setup_view_drawings_tab(self):
        self.tab_drawings.columnconfigure(0, weight=1)
        self.create_section_header("VIEW SIGNED DRAWINGS (SERVER)", row=0, parent=self.tab_drawings)
        
        self.search_draw_frame = ctk.CTkFrame(self.tab_drawings, fg_color="transparent")
        self.search_draw_frame.grid(row=1, column=0, padx=40, pady=40, sticky="ew")
        self.search_draw_frame.columnconfigure(0, weight=1)

        ctk.CTkLabel(self.search_draw_frame, text="Enter Part Number to Search Signed Copy:", font=FONT_LABEL).grid(row=0, column=0, sticky="w", pady=(0, 10))
        self.pdf_search_ent = ctk.CTkEntry(self.search_draw_frame, font=("Segoe UI", 16), height=45, placeholder_text="e.g. H10-100-001")
        self.pdf_search_ent.grid(row=1, column=0, sticky="ew", padx=(0, 15))
        self.pdf_search_ent.bind("<KeyRelease>", lambda e: self.to_uppercase(e, self.pdf_search_ent))
        
        self.pdf_search_ent.bind("<Return>", lambda e: self.search_pdf_thread())

        self.btn_search_pdf = ctk.CTkButton(self.search_draw_frame, text="SEARCH DRAWING", fg_color=COLOR_ACCENT, 
                                           height=45, width=200, font=("Segoe UI", 13, "bold"), command=self.search_pdf_thread)
        self.btn_search_pdf.grid(row=1, column=1)

        self.pdf_status_lbl = ctk.CTkLabel(self.tab_drawings, text="", font=("Segoe UI", 12))
        self.pdf_status_lbl.grid(row=2, column=0, pady=10)


    # =========================================================================
    # LOGIC FUNCTIONS DARI CODE 1 (DATA ENTRY)
    # =========================================================================
    def create_section_header(self, text, row, parent):
        label = ctk.CTkLabel(parent, text=text, font=FONT_SECTION, text_color=COLOR_PRIMARY, anchor="w")
        label.grid(row=row, column=0, columnspan=2, padx=20, pady=(15, 5), sticky="w")
        line = ctk.CTkFrame(parent, height=2, fg_color="#E0E0E0")
        line.grid(row=row+1, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="ew")

    def add_input_field(self, label, row, col, parent):
        ctk.CTkLabel(parent, text=label, font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=row, column=col, padx=20, pady=(5, 0), sticky="w")

    def clear_field_error(self, field):
        if field == "draw": self.err_draw.configure(text="")
        elif field == "part": self.err_part.configure(text="")
        elif field == "rev": self.err_rev.configure(text="")
        elif field == "total": self.err_total.configure(text="")

    def update_status(self, m, c=COLOR_TEXT): 
        self.after(0, lambda: self.status_label.configure(text=f"STATUS: {m}", text_color=c))

    def trigger_feedback(self, t="success", m="Saved Successfully"):
        c = COLOR_SUCCESS if t == "success" else COLOR_DANGER if t == "error" else COLOR_INFO
        self.lbl_feedback.configure(text=m, fg_color=c)
        self.after(5000, lambda: self.lbl_feedback.configure(text="", fg_color="transparent"))

    def set_process_highlight(self, message, color=COLOR_INFO):
        self.lbl_feedback.configure(text=message, fg_color=color)

    def open_folder(self):
        if os.path.exists(BASE_PATH): os.startfile(BASE_PATH)
        else: messagebox.showerror("Error", "Folder not found")

    def open_project_file(self):
        fp = os.path.join(BASE_PATH, f"{self.proj_v.get()}.xlsx")
        if os.path.exists(fp): os.startfile(fp)
        else: messagebox.showerror("Error", "File not found")

    def to_uppercase(self, event, widget):
        if event.keysym in ["Control_L", "Control_R", "Shift_L", "Shift_R", "Left", "Right"]: return
        pos = widget.index(ctk.INSERT); val = widget.get().upper()
        if widget.get() != val: widget.delete(0, 'end'); widget.insert(0, val); widget.icursor(pos)

    def on_part_input(self, event):
        self.to_uppercase(event, self.part_ent)
        if self.proj_v.get() != "Wheel Press Machine" and "(N)" in self.part_ent.get(): self.batch_v.set("N")

    def auto_remark_logic(self, event):
        val = self.rev_ent.get()
        if val.isdigit(): self.remark_v.set("Revised" if int(val) >= 1 else "New")

    def update_logic(self, choice):
        self.batch_drop.configure(state="normal")
        if choice == "Wheel Press Machine":
            self.batch_v.set("-"); self.batch_drop.configure(state="disabled") 
            self.assembly_drop.configure(values=["Wheel Press"]); self.assembly_v.set("Wheel Press")
            self.eng_v.set("Baskaran"); self.eng_drop.configure(values=["Baskaran"])
        elif choice in ["M10", "N10"]:
            self.batch_v.set("-"); self.assembly_drop.configure(values=ASSEMBLIES + ["Auxiliary Hood"])
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)
        elif choice == "H10 TRC":
            self.batch_v.set("2"); self.assembly_drop.configure(values=ASSEMBLIES)
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)
        else:
            self.assembly_drop.configure(values=ASSEMBLIES)
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)

    def set_today(self): self.date_picker.set_date(date.today())

    def clear_all(self):
        self.update_logic(self.proj_v.get())
        for w in [self.draw_ent, self.part_ent, self.rev_ent, self.total_ent]: w.delete(0, 'end')
        self.txt_reason.delete("1.0", "end") # Clear textbox
        self.err_draw.configure(text=""); self.err_part.configure(text=""); self.err_rev.configure(text=""); self.err_total.configure(text="")
        self.lbl_feedback.configure(text="", fg_color="transparent")

    def check_part_existence_thread(self):
        self.btn_check_part.configure(state="disabled", text="...")
        self.set_process_highlight("🔍 SEARCHING DATABASE & FILES... PLEASE WAIT", COLOR_WARNING)
        threading.Thread(target=self._run_global_search, args=(self.part_ent.get().strip().upper(),), daemon=True).start()

    def _run_global_search(self, part):
        if not part: self.after(0, lambda: self._post_check([])); return
        matches = []
        m_path = os.path.join(BASE_PATH, MASTER_FILE)
        try:
            if os.path.exists(m_path):
                wb = load_workbook(m_path, read_only=True); ws = wb.active
                for row in ws.iter_rows(min_row=2, max_row=8000, values_only=True): 
                    if row and len(row) >= 9 and str(row[5]).strip().upper() == part:
                        matches.append({'project': row[0], 'country': row[1], 'batch': row[2], 'assembly': row[3], 'draw': row[4], 'part': row[5], 'rev': row[6], 'total': row[7], 'eng': row[8]})
                wb.close()
            for ui_proj in PROJECTS:
                sc = PROJ_MAP.get(ui_proj); p_f = os.path.join(BASE_PATH, f"{ui_proj}.xlsx")
                if os.path.exists(p_f):
                    wb = load_workbook(p_f, read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=3, values_only=True): 
                            if row and len(row) >= 5 and str(row[2]).strip().upper() == part:
                                if not any(m['project'] == sc and str(m['rev']) == str(row[3]) for m in matches):
                                    matches.append({'project': sc, 'country': "Tanzania" if "TRC" in ui_proj else "Malaysia", 'batch': "-", 'assembly': sheet_name, 'draw': row[1], 'part': row[2], 'rev': row[3], 'total': row[4], 'eng': "-"})
                    wb.close()
        except: pass
        
        matches.sort(key=lambda m: (try_int(m['rev'])), reverse=True)
        self.after(0, lambda: self._post_check(matches, part))

    def _post_check(self, matches, part_searched=None):
        self.btn_check_part.configure(state="normal", text="Search")
        self.lbl_feedback.configure(text="", fg_color="transparent") 
        if not matches:
             if part_searched: self.trigger_feedback("info", "Part Not Found")
             return
        unique_results = []
        seen_keys = set()
        for m in matches:
            key = (m['project'], m['draw'])
            if key not in seen_keys: seen_keys.add(key); unique_results.append(m)
        if len(unique_results) == 1: self.show_match_found_dialog(unique_results[0])
        else: self.open_selection_window(unique_results)

    def show_match_found_dialog(self, m):
        top = ctk.CTkToplevel(self); top.title("Record Found"); top.geometry("650x450"); top.transient(self); top.grab_set()
        ctk.CTkLabel(top, text="EXISTING RECORD FOUND", font=("Segoe UI", 16, "bold"), text_color=COLOR_SUCCESS).pack(pady=15)
        info_frame = ctk.CTkFrame(top, fg_color="white", border_width=2, border_color="#BDC3C7", corner_radius=10); info_frame.pack(fill="x", padx=20, pady=10)
        
        def add_row_info(r, lbl, val, is_bold=True, wrap=0):
            ctk.CTkLabel(info_frame, text=lbl, font=("Segoe UI", 12), text_color="gray").grid(row=r, column=0, padx=15, pady=5, sticky="w")
            label_val = ctk.CTkLabel(info_frame, text=str(val), font=("Segoe UI", 14, "bold" if is_bold else "normal"), text_color=COLOR_ACCENT, wraplength=wrap if wrap > 0 else 0, justify="left", anchor="w")
            label_val.grid(row=r, column=1, padx=5, pady=5, sticky="w")

        add_row_info(0, "Project:", m['project'])
        add_row_info(1, "Assembly:", m['assembly'])
        add_row_info(2, "Drawing:", m['draw'], wrap=450)
        add_row_info(3, "Revision:", m['rev'])
        add_row_info(4, "Part No:", m['part'])

        ctk.CTkButton(top, text="YES (Auto-Fill)", fg_color=COLOR_SUCCESS, height=40, font=("Segoe UI", 12, "bold"), command=lambda: [self.autofill_form(m), top.destroy()]).pack(pady=10)
        ctk.CTkButton(top, text="CANCEL", fg_color=COLOR_DANGER, width=100, command=top.destroy).pack(pady=5)

    def open_selection_window(self, matches):
        top = ctk.CTkToplevel(self); top.title("Search Results"); top.geometry("850x600"); top.transient(self); top.grab_set()
        h = ctk.CTkFrame(top, fg_color=COLOR_PRIMARY, height=70, corner_radius=0); h.pack(fill="x")
        ctk.CTkLabel(h, text="SELECT DATA (Sorted by Latest Revision)", font=("Segoe UI", 18, "bold"), text_color="white").pack(pady=15)
        s = ctk.CTkScrollableFrame(top, fg_color="transparent"); s.pack(fill="both", expand=True, padx=20, pady=10)
        
        for i, m in enumerate(matches):
            is_latest = (i == 0) 
            card_border_color = COLOR_SUCCESS if is_latest else "#BDC3C7"
            card_border_width = 3 if is_latest else 1
            
            c = ctk.CTkFrame(s, fg_color="white", border_width=card_border_width, border_color=card_border_color, corner_radius=10)
            c.pack(fill="x", pady=8, padx=5, ipady=5)
            
            content_f = ctk.CTkFrame(c, fg_color="transparent")
            content_f.pack(side="left", padx=15, pady=10, fill="both", expand=True)
            
            if is_latest:
                badge_f = ctk.CTkFrame(content_f, fg_color=COLOR_SUCCESS, corner_radius=5)
                badge_f.pack(anchor="w", pady=(0, 5))
                ctk.CTkLabel(badge_f, text=" LATEST REVISION ", font=("Segoe UI", 10, "bold"), text_color="white").pack(padx=5, pady=2)
            
            ctk.CTkLabel(content_f, text=f"PROJECT: {m['project']} ({m.get('country', 'N/A')})", font=("Segoe UI", 13, "bold"), text_color=COLOR_ACCENT, anchor="w").pack(fill="x")
            ctk.CTkLabel(content_f, text=f"DRAWING: {m['draw']}", font=("Segoe UI", 12, "bold"), text_color=COLOR_PRIMARY, wraplength=550, justify="left", anchor="w").pack(fill="x", pady=(2, 0))
            ctk.CTkLabel(content_f, text=f"Assy: {m['assembly']}   |   Part: {m['part']}   |   Rev: {m['rev']}", font=("Segoe UI", 11), text_color="gray", anchor="w").pack(fill="x", pady=(2, 0))
            ctk.CTkButton(c, text="SELECT", width=100, height=40, fg_color=COLOR_SUCCESS, font=("Segoe UI", 11, "bold"), command=lambda data=m: [self.autofill_form(data), top.destroy()]).pack(side="right", padx=15)
            
        ctk.CTkButton(top, text="CANCEL / CLOSE", fg_color=COLOR_DANGER, width=150, height=40, font=("Segoe UI", 12, "bold"), command=top.destroy).pack(pady=20)

    def autofill_form(self, data):
        sc = str(data.get('project')).strip()
        ci = str(data.get('country', '')).upper()
        ui = "H10 TRC" if sc == "H10" and "TANZANIA" in ci else "H10 BeraPit" if sc == "H10" else REVERSE_PROJ_MAP.get(sc)
        if ui: self.proj_v.set(ui); self.update_logic(ui) 
        if data.get('batch'): self.batch_v.set("-" if str(data.get('batch')) in [None, "", "None"] else str(data.get('batch')))
        if data.get('assembly'): self.assembly_v.set(str(data.get('assembly')))
        if data.get('draw'): self.draw_ent.delete(0, 'end'); self.draw_ent.insert(0, str(data.get('draw')))
        if data.get('rev') is not None: self.rev_ent.delete(0, 'end'); self.rev_ent.insert(0, str(data.get('rev'))); self.auto_remark_logic(None)
        if data.get('total') is not None: self.total_ent.delete(0, 'end'); self.total_ent.insert(0, str(data.get('total')))
        if data.get('eng') and data.get('eng') in ENGINEER_LIST: self.eng_v.set(data.get('eng'))

    def check_duplicate_entry(self, fp, tsn, pi, ri):
        if not os.path.exists(fp): return False 
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
            sn = next((s for s in wb.sheetnames if s.strip().lower() == tsn.lower()), None)
            if not sn: wb.close(); return False 
            ws = wb[sn]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and len(row) >= 4 and str(row[2]).strip().upper() == pi and str(row[3]) == str(ri):
                    wb.close(); return True
            wb.close(); return False
        except: return False 

    def check_master_duplicate(self, cp):
        mp = os.path.join(BASE_PATH, MASTER_FILE)
        if not os.path.exists(mp): return False
        try:
            wb = load_workbook(mp, read_only=True, data_only=True); ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=8000, values_only=True):
                if not row or len(row) < 9: continue
                match = True
                for i in range(9): 
                    if str(row[i]).strip().upper() != str(cp[i]).strip().upper():
                        match = False; break
                if match: wb.close(); return True 
            wb.close(); return False
        except: return False

    def show_duplicate_choice_dialog(self, part, rev, sql_data, excel_master_data, master_file_path):
        top = ctk.CTkToplevel(self); top.title("Action Required"); top.geometry("480x320"); top.transient(self); top.grab_set()
        ctk.CTkLabel(top, text="DUPLICATE IN PROJECT EXCEL", font=FONT_SECTION, text_color=COLOR_DANGER).pack(pady=15)
        ctk.CTkLabel(top, text=f"Part No: {part} (Rev {rev}) already exists in sub-project.").pack(pady=5)
        btn_f = ctk.CTkFrame(top, fg_color="transparent"); btn_f.pack(pady=15)
        
        def h_master():
            top.destroy()
            try:
                if not os.path.exists(master_file_path):
                    wb_m = Workbook(); ws_m = wb_m.active; ws_m.title = "MasterList"; ws_m.append(MASTER_HEADERS); wb_m.save(master_file_path)
                wb_m = load_workbook(master_file_path); ws_m = wb_m.active
                re_release_data = list(excel_master_data); re_release_data[10] = "Re-Release" 
                l_row = ws_m.max_row
                for idx in range(1, 12):
                    c = ws_m.cell(row=l_row, column=idx)
                    if idx == 10: c.number_format = 'DD/MM/YYYY'
                    c.alignment = LEFT_ALIGN if idx in [4, 5, 6] else CENTER_ALIGN; c.border = THIN_BORDER
                wb_m.save(master_file_path); self.trigger_feedback("success", "Logged to Master Only")
            except Exception as e: messagebox.showerror("Error", str(e))
            
        def h_sql():
            top.destroy()
            try:
                conn = psycopg2.connect(**DB_CONFIG); cur = conn.cursor()
                cur.execute("""SELECT 1 FROM project_data WHERE project_name=%s AND country=%s AND batch=%s AND main_assembly=%s AND drawing_name=%s AND part_number=%s AND revision=%s AND total_sheets=%s AND engineer=%s""", 
                            (sql_data[0], sql_data[1], sql_data[2], sql_data[3], sql_data[4], sql_data[5], sql_data[6], sql_data[7], sql_data[8]))
                if not cur.fetchone():
                    cur.execute("INSERT INTO project_data (project_name, country, batch, main_assembly, drawing_name, part_number, revision, total_sheets, engineer, date_approved, remarks) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", sql_data)
                    conn.commit(); self.trigger_feedback("success", "SQL Only Inserted")
                else: self.update_status("Already in SQL", COLOR_WARNING)
                cur.close(); conn.close()
            except Exception as e: messagebox.showerror("DB Error", str(e))
            
        ctk.CTkButton(btn_f, text="RE-RELEASE\n(MASTER ONLY)", fg_color=COLOR_PRIMARY, width=150, height=50, command=h_master).grid(row=0, column=0, padx=10)
        ctk.CTkButton(btn_f, text="SQL ONLY\n(DATABASE)", fg_color=COLOR_INFO, width=150, height=50, command=h_sql).grid(row=0, column=1, padx=10)
        ctk.CTkButton(top, text="CANCEL", fg_color=COLOR_DANGER, width=80, command=top.destroy).pack(pady=15)

    def submit(self):
        self.err_draw.configure(text="")
        self.err_part.configure(text="")
        self.err_rev.configure(text="")
        self.err_total.configure(text="")
        full_name = self.proj_v.get(); short_proj = PROJ_MAP.get(full_name)
        proj_file = os.path.join(BASE_PATH, f"{full_name}.xlsx"); master_file_path = os.path.join(BASE_PATH, MASTER_FILE)
        draw = self.draw_ent.get().upper().strip(); part = self.part_ent.get().upper().strip()
        rev_str = self.rev_ent.get().strip(); total_str = self.total_ent.get().strip()
        
        reason_val = self.txt_reason.get("1.0", "end-1c").strip()
        
        has_err = False
        if not draw: self.err_draw.configure(text="* Sila isi Drawing Name"); has_err = True
        if not part: self.err_part.configure(text="* Sila isi Part Number"); has_err = True
        if not rev_str: self.err_rev.configure(text="* Sila isi Revision"); has_err = True
        if not total_str: self.err_total.configure(text="* Sila isi Total Sheets"); has_err = True
        if has_err: self.trigger_feedback("error", "Isi ruangan bertanda merah!"); return
        
        if self.remark_v.get() == "Revised" and not reason_val:
            self.trigger_feedback("error", "Reason for Revision is compulsory!")
            return
        
        try: rev, total = int(rev_str), int(total_str)
        except: self.trigger_feedback("error", "Revision/Total must be numbers"); return

        self.set_process_highlight("💾 PROCESSING DATA... PLEASE DO NOT CLOSE", COLOR_WARNING)
        tsn = self.assembly_v.get()[:31]; batch_val = self.batch_v.get(); bve = int(batch_val) if batch_val.isdigit() else batch_val
        dt_obj = self.date_picker.get_date(); dt_sql = dt_obj.strftime('%Y-%m-%d'); country = "Tanzania" if full_name == "H10 TRC" else "Malaysia"

        for f in [proj_file, master_file_path]:
            if os.path.exists(f):
                try: 
                    with open(f, "a"): pass
                except PermissionError: 
                    self.lbl_feedback.configure(text="", fg_color="transparent")
                    messagebox.showerror("Error", f"Close {os.path.basename(f)}!"); return

        sql_data = [short_proj, country, batch_val, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get(), dt_sql, self.remark_v.get()]
        excel_master_data = [short_proj, country, bve, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get(), dt_obj, self.remark_v.get()]

        if self.check_duplicate_entry(proj_file, tsn, part, rev):
            self.lbl_feedback.configure(text="", fg_color="transparent")
            self.show_duplicate_choice_dialog(part, rev, sql_data, excel_master_data, master_file_path); return

        try:
            conn = psycopg2.connect(**DB_CONFIG); cur = conn.cursor()
            cur.execute("""SELECT 1 FROM project_data WHERE project_name=%s AND country=%s AND batch=%s AND main_assembly=%s AND drawing_name=%s AND part_number=%s AND revision=%s AND total_sheets=%s AND engineer=%s""", 
                        (short_proj, country, batch_val, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get()))
            if not cur.fetchone():
                cur.execute("INSERT INTO project_data (project_name, country, batch, main_assembly, drawing_name, part_number, revision, total_sheets, engineer, date_approved, remarks) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", sql_data)
                conn.commit()
            cur.close(); conn.close()

            if not os.path.exists(master_file_path):
                wb_m = Workbook(); ws_m = wb_m.active; ws_m.title = "MasterList"; ws_m.append(MASTER_HEADERS); wb_m.save(master_file_path)
            add_to_master = True; final_master_data = list(excel_master_data)
            if self.check_master_duplicate(excel_master_data):
                self.lbl_feedback.configure(text="", fg_color="transparent")
                if messagebox.askyesno("Master Log", "Matches Master exactly. Re-Release?"): final_master_data[10] = "Re-Release"
                else: add_to_master = False
                self.set_process_highlight("💾 CONTINUING WRITE PROCESS...", COLOR_WARNING)
            if add_to_master:
                wb_m = load_workbook(master_file_path); ws_m = wb_m.active; ws_m.append(final_master_data); lr = ws_m.max_row
                for idx in range(1, 12):
                    cell = ws_m.cell(row=lr, column=idx)
                    if idx == 10: cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = LEFT_ALIGN if idx in [4, 5, 6] else CENTER_ALIGN
                    cell.border = THIN_BORDER
                wb_m.save(master_file_path)

            wb_p = load_workbook(proj_file)
            asn = next((s for s in wb_p.sheetnames if s.strip().lower() == tsn.lower()), None)
            if not asn:
                ws_p = wb_p.create_sheet(tsn); ws_p.append(["Document Control"]); ws_p.append(["Sl. No", "Drawing Name", "Part Number", "Revision", "Total Drawings", "Date Approved", "Remarks"])
                ws_p.cell(row=15, column=1, value="Drawing issued by:-"); dsr = 3
            else:
                ws_p = wb_p[asn]; dsr = 3
                for r in range(1, 11):
                    if "part number" in str(ws_p.cell(row=r, column=3).value or "").lower(): dsr = r + 1; break
            sig = next((r for r in range(dsr, ws_p.max_row + 50) if any("issued by" in str(ws_p.cell(row=r, column=c).value or "").lower() for c in range(1, 6))), None)
            if not sig: sig = max(ws_p.max_row + 2, 4); ws_p.cell(row=sig, column=1, value="Drawing issued by:-")
            tr = next((r for r in range(dsr, sig) if str(ws_p.cell(row=r, column=3).value or "").strip().upper() == part), None)
            is_ov = tr is not None; rts = "Revised" if is_ov else self.remark_v.get()
            if not tr:
                ldr = dsr - 1
                for r in range(max(dsr, sig - 1), dsr - 1, -1):
                    if ws_p.cell(row=r, column=2).value or ws_p.cell(row=r, column=3).value: ldr = r; break
                tr = ldr + 1; ws_p.insert_rows(tr)
                for r in range(ws_p.max_row, tr, -1):
                    if (r-1) in ws_p.row_dimensions: ws_p.row_dimensions[r].height = ws_p.row_dimensions[r-1].height
            sl_no = tr - dsr + 1 if not is_ov else ws_p.cell(row=tr, column=1).value
            fd = [sl_no, draw, part, rev, total, dt_obj, rts]
            for col, val in enumerate(fd, 1):
                cell = ws_p.cell(row=tr, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = val
                    if col == 6: cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = LEFT_ALIGN if col in [2, 3] else CENTER_ALIGN
                    cell.border = THIN_BORDER
            sig_rc = next((r for r in range(dsr, ws_p.max_row + 10) if any("issued by" in str(ws_p.cell(row=r, column=c).value or "").lower() for c in range(1, 6))), ws_p.max_row + 1)
            dm = []
            if sig_rc - 1 >= dsr:
                for r in range(dsr, sig_rc):
                    rd = [ws_p.cell(row=r, column=c).value for c in range(2, 8)]
                    if any(rd): dm.append(rd)
                dm.sort(key=lambda x: str(x[1]).strip().upper() if x[1] else "")
                for i, rd in enumerate(dm):
                    cr = dsr + i; slc = ws_p.cell(row=cr, column=1); slc.value = i+1; slc.alignment = CENTER_ALIGN; slc.border = THIN_BORDER
                    for idx, v in enumerate(rd):
                        c = idx + 2; cell = ws_p.cell(row=cr, column=c); cell.value = v
                        if c == 6: cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = LEFT_ALIGN if c in [2, 3] else CENTER_ALIGN
                        cell.border = THIN_BORDER
            dob = []
            for r in range(dsr, sig_rc):
                dv = ws_p.cell(row=r, column=6).value
                if isinstance(dv, (datetime, date)): dob.append(dv.date() if isinstance(dv, datetime) else dv)
            ldo = max(dob) if dob else None
            for r in range(dsr, sig_rc):
                dv = ws_p.cell(row=r, column=6).value
                cdo = dv.date() if isinstance(dv, datetime) else dv if isinstance(dv, date) else None
                isl = (cdo == ldo and ldo is not None)
                for c in range(1, 8):
                    cell = ws_p.cell(row=r, column=c)
                    if isinstance(cell, MergedCell): continue
                    if c == 7 and not isl: cell.value = None
                    cell.fill = CREAM_YELLOW if isl else NO_FILL
                    cell.font = Font(bold=isl)
                    ws_p.row_dimensions[r].height = 18
            wb_p.save(proj_file)
            
            # --- SYNC KE DC TRANSMITTAL LIST ---
            self.add_to_dc_list_excel(full_name, draw, part, rev, total, dt_obj, self.assembly_v.get(), reason_val)
            
            # PELARASAN: Tiada auto clear dipanggil
            self.trigger_feedback("success", "Saved & Synced!")
        except Exception as e: messagebox.showerror("Error", f"Submit Error: {e}")

    def add_to_dc_list_excel(self, proj, draw, part, rev, total, appr_date, assy, reason_val):
        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj] if proj in wb.sheetnames else wb.create_sheet(proj)
            last_row = ws.max_row + 1
            sl = 1 if last_row == 2 else (ws.cell(row=last_row-1, column=1).value or 0) + 1
            
            # Insert into column M (13)
            row_data = [sl, draw, part, rev, total, appr_date, assy, "", "", "", "", "", reason_val]
            ws.append(row_data)
            
            wb.save(DC_LIST_FILE)
            wb.close()
        except: pass

    # =========================================================================
    # TAB 2: SEARCH PDF LOGIC (DARI KOD 1 ASAL)
    # =========================================================================
    def search_pdf_thread(self):
        part_no = self.pdf_search_ent.get().strip().upper()
        if not part_no: return
        self.btn_search_pdf.configure(state="disabled", text="Searching...")
        self.pdf_status_lbl.configure(text="Searching...", text_color=COLOR_PRIMARY)
        threading.Thread(target=self._run_pdf_search_dynamic, args=(part_no,), daemon=True).start()

    def _run_pdf_search_dynamic(self, part_no):
        found_matches = []
        master_records = []
        m_path = os.path.join(BASE_PATH, MASTER_FILE)
        try:
            if os.path.exists(m_path):
                wb_m = load_workbook(m_path, data_only=True, read_only=True)
                ws_m = wb_m.active
                for row in ws_m.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 9 and str(row[5]).strip().upper() == part_no:
                        master_records.append({"project": row[0], "assy": row[3], "draw": row[4], "part": row[5], "rev": row[6], "total": row[7]})
                wb_m.close()
        except: pass
        
        if not os.path.exists(BASE_DRAWINGS_PATH):
            self.after(0, lambda: self.pdf_status_lbl.configure(text="❌ Drive Y: unreachable.", text_color=COLOR_DANGER))
            self.after(0, lambda: self.btn_search_pdf.configure(state="normal", text="SEARCH DRAWING"))
            return
            
        try:
            for proj_folder in os.listdir(BASE_DRAWINGS_PATH):
                proj_path = os.path.join(BASE_DRAWINGS_PATH, proj_folder)
                if not os.path.isdir(proj_path): continue
                for assy_folder in os.listdir(proj_path):
                    assy_path = os.path.join(proj_path, assy_folder)
                    if not os.path.isdir(assy_path): continue
                    
                    latest_file = os.path.join(assy_path, f"{part_no}.pdf")
                    if os.path.exists(latest_file):
                        meta = self._find_best_meta(part_no, proj_folder, assy_folder, master_records, is_latest=True)
                        found_matches.append({**meta, "path": latest_file, "type": "Latest"})
                    
                    for item in os.listdir(assy_path):
                        item_path = os.path.join(assy_path, item)
                        if os.path.isdir(item_path) and "obsolete" in item.lower():
                            for obs_file in os.listdir(item_path):
                                if obs_file.upper().startswith(part_no) and obs_file.lower().endswith(".pdf"):
                                    rev_f = self._extract_rev_from_filename(obs_file)
                                    meta = self._find_best_meta(part_no, proj_folder, assy_folder, master_records, rev_to_match=rev_f)
                                    found_matches.append({**meta, "path": os.path.join(item_path, obs_file), "type": "Obsolete"})
        except: pass
        self.after(0, lambda: self._handle_pdf_results(found_matches))

    def _find_best_meta(self, part, proj, assy, records, is_latest=False, rev_to_match=None):
        default = {"part": part, "draw": "Unknown Title", "project": proj, "assy": assy, "rev": "N/A", "total": "N/A"}
        short_proj = PROJ_MAP.get(proj, proj)
        matches = [r for r in records if r["project"] == short_proj]
        if not matches: return default
        if is_latest:
            matches.sort(key=lambda x: try_int(x["rev"]), reverse=True)
            return matches[0]
        elif rev_to_match is not None:
            for r in matches:
                if str(r["rev"]) == str(rev_to_match): return r
        return matches[0]

    def _extract_rev_from_filename(self, filename):
        try:
            if "REV" in filename.upper(): return filename.upper().split("REV")[-1].strip().split(".")[0].strip()
        except: pass
        return None

    def _handle_pdf_results(self, results):
        self.btn_search_pdf.configure(state="normal", text="SEARCH DRAWING")
        if not results:
            self.pdf_status_lbl.configure(text="❌ No record found on server.", text_color=COLOR_DANGER)
            messagebox.showwarning("No Records Found", f"No signed drawings found for {self.pdf_search_ent.get()} on server.")
            return
        self.pdf_status_lbl.configure(text=f"✅ Found {len(results)} drawing(s).", text_color=COLOR_SUCCESS)
        self.show_pdf_selection_dialog(results)

    def show_pdf_selection_dialog(self, results):
        top = ctk.CTkToplevel(self)
        top.title("Drawing Records Found")
        top.geometry("850x600")
        top.transient(self)
        top.grab_set()
        
        ctk.CTkLabel(top, text="SIGNED DRAWINGS FOUND", font=("Segoe UI", 18, "bold"), text_color=COLOR_ACCENT).pack(pady=15)
        ctk.CTkLabel(top, text="Select a record to open the Signed Copy PDF:", font=("Segoe UI", 12)).pack(pady=5)

        scroll = ctk.CTkScrollableFrame(top, width=780, height=400, fg_color="transparent")
        scroll.pack(padx=20, pady=10, fill="both", expand=True)

        for res in results:
            is_latest = res["type"] == "Latest"
            card_border_color = COLOR_SUCCESS if is_latest else "#BDC3C7"
            card_border_width = 3 if is_latest else 1
            
            card = ctk.CTkFrame(scroll, fg_color="white", border_width=card_border_width, border_color=card_border_color, corner_radius=10)
            card.pack(fill="x", pady=8, padx=5)
            
            main_f = ctk.CTkFrame(card, fg_color="transparent")
            main_f.pack(side="left", padx=15, pady=12, fill="both", expand=True)
            
            if is_latest:
                badge_f = ctk.CTkFrame(main_f, fg_color=COLOR_SUCCESS, corner_radius=5)
                badge_f.pack(anchor="w", pady=(0, 5))
                ctk.CTkLabel(badge_f, text=" LATEST REVISION ", font=("Segoe UI", 10, "bold"), text_color="white").pack(padx=5, pady=2)

            row1 = ctk.CTkFrame(main_f, fg_color="transparent")
            row1.pack(fill="x")
            ctk.CTkLabel(row1, text=res["part"], font=("Segoe UI", 14, "bold"), text_color=COLOR_PRIMARY).pack(side="left")
            title_lbl = ctk.CTkLabel(row1, text=f"  |  {res['draw']}", font=("Segoe UI", 13), text_color=COLOR_TEXT, wraplength=450, justify="left", anchor="w")
            title_lbl.pack(side="left", fill="x", expand=True)

            row2 = ctk.CTkFrame(main_f, fg_color="transparent")
            row2.pack(fill="x", pady=(2, 0))
            ctk.CTkLabel(row2, text=f"Project: {res['project']}   •   Assembly: {res['assy']}", font=("Segoe UI", 11), text_color="gray").pack(side="left")

            row3 = ctk.CTkFrame(main_f, fg_color="transparent")
            row3.pack(fill="x")
            ctk.CTkLabel(row3, text=f"Revision: {res['rev']}   •   Sheets: {res['total']}", font=("Segoe UI", 11, "bold"), text_color=COLOR_ACCENT).pack(side="left")

            btn_color = COLOR_SUCCESS if is_latest else "#95a5a6"
            ctk.CTkButton(card, text="OPEN PDF", width=110, height=40, fg_color=btn_color, font=("Segoe UI", 12, "bold"), 
                          command=lambda p=res["path"]: os.startfile(p)).pack(side="right", padx=15)

        ctk.CTkButton(top, text="CLOSE WINDOW", fg_color=COLOR_DANGER, width=150, height=40, font=("Segoe UI", 12, "bold"), command=top.destroy).pack(pady=20)


    def publish_to_server_thread(self):
        self.btn_publish.configure(state="disabled", text="Publishing...")
        threading.Thread(target=self._run_publish_logic, daemon=True).start()

    def _run_publish_logic(self):
        if not os.path.exists(SERVER_PATH):
            self.after(0, lambda: messagebox.showerror("Error", "Server unreachable")); return
        suc, fail = 0, 0
        for l, s in FILE_MAPPING.items():
            lf, sf = os.path.join(BASE_PATH, l), os.path.join(SERVER_PATH, s)
            if not os.path.exists(lf): fail += 1; continue
            try:
                if os.path.exists(sf): os.rename(sf, sf)
                shutil.copy2(lf, sf); suc += 1
            except: fail += 1
        self.after(0, self._finalize_publish, suc, fail)

    def _finalize_publish(self, s, f):
        self.lbl_last_sync.configure(text=f"Last Sync: {datetime.now().strftime('%H:%M:%S')}")
        self.btn_publish.configure(state="normal", text="PUBLISH TO SERVER")
        messagebox.showinfo("Result", f"Success: {s} | Fail: {f}")


    # =========================================================================
    # TAB 3: DRAWING TRANSMITTAL (Sourced exactly from Code 2)
    # =========================================================================
    def setup_hannan_dashboard(self):
        top_bar = ctk.CTkFrame(self.tab_trans, fg_color="transparent")
        top_bar.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(top_bar, text="Engineering Transmittal Dashboard", font=("Segoe UI", 20, "bold"), text_color="#2C3E50").pack(side="left")
        ctk.CTkButton(top_bar, text="REFRESH ALL", font=("Segoe UI", 13, "bold"), command=self.refresh_hannan_data).pack(side="right")

        self.trans_tabview = ctk.CTkTabview(self.tab_trans, corner_radius=15, segmented_button_selected_color="#2C3E50")
        self.trans_tabview.pack(fill="both", expand=True)
        
        self.project_scrolls = {}
        for proj in PROJECTS:
            tab = self.trans_tabview.add(proj)
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(1, weight=1)
            
            info_f = ctk.CTkFrame(tab, fg_color="transparent")
            info_f.grid(row=0, column=0, sticky="ew", pady=10, padx=15)
            
            self.project_scrolls[proj] = {
                "scroll": ctk.CTkScrollableFrame(tab, fg_color="#FBFCFC"),
                "note_label": ctk.CTkLabel(info_f, text="", font=("Segoe UI", 12, "bold")),
                "issue_btn": ctk.CTkButton(info_f, text="Create & Issue (All)", fg_color="#27AE60", font=("Segoe UI", 13, "bold"), command=lambda p=proj: self.create_issue_action(p))
            }
            self.project_scrolls[proj]["note_label"].pack(side="left")
            self.project_scrolls[proj]["issue_btn"].pack(side="right")
            self.project_scrolls[proj]["scroll"].grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

        self.refresh_hannan_data()

    def refresh_hannan_data(self):
        if not os.path.exists(DC_LIST_FILE): return
        try:
            wb = load_workbook(DC_LIST_FILE, data_only=True)
            now_dt = datetime.now()
            
            for proj in PROJECTS:
                if proj not in self.project_scrolls: continue
                scroll = self.project_scrolls[proj]["scroll"]
                note = self.project_scrolls[proj]["note_label"]
                
                for w in scroll.winfo_children(): w.destroy()
                
                if proj not in wb.sheetnames: continue
                ws = wb[proj]
                
                raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
                rows = []
                for r in raw_rows:
                    if r is None or not r[1]: continue
                    r_list = list(r)
                    while len(r_list) < 14: r_list.append("")
                    rows.append(r_list)
                
                unissued_groups = {} 
                pending_groups = {}  
                received_groups = {} 

                for r in rows:
                    assy = str(r[6])
                    iss_time = str(r[8]) 
                    f_id = str(r[11])    
                    
                    if not r[7]: 
                        if assy not in unissued_groups: unissued_groups[assy] = []
                        unissued_groups[assy].append(r)
                    elif not r[9]: 
                        key = (assy, iss_time)
                        if key not in pending_groups: pending_groups[key] = []
                        pending_groups[key].append(r)
                    else: 
                        rec_time_val = r[10]
                        try:
                            if isinstance(rec_time_val, str): rec_dt = datetime.strptime(rec_time_val, "%Y-%m-%d %H:%M:%S")
                            else: rec_dt = rec_time_val
                            if rec_dt and (now_dt - rec_dt) < timedelta(hours=1):
                                if f_id not in received_groups: received_groups[f_id] = []
                                received_groups[f_id].append(r)
                        except: pass

                unissued_count = len(unissued_groups.keys())
                pending_count = len(pending_groups.keys())
                note.configure(text=f"Unissued Groups: {unissued_count}  |  Pending DC: {pending_count} Batches", 
                               text_color="#C0392B" if unissued_count > 0 else "#27AE60")
                
                if unissued_count > 0:
                    self.trans_tabview._segmented_button._buttons_dict[proj].configure(text_color="#C0392B")
                else:
                    self.trans_tabview._segmented_button._buttons_dict[proj].configure(text_color="#2C3E50")

                self.draw_hannan_header(scroll)

                for assy, items in unissued_groups.items():
                    self.draw_hannan_group_row(scroll, proj, assy, items, "READY", "#3498DB")

                for (assy, time), items in pending_groups.items():
                    display_label = f"{assy} (Issued: {time})"
                    self.draw_hannan_group_row(scroll, proj, display_label, items, "ISSUED", "#E67E22")

                for fid, items in received_groups.items():
                    assy_name = items[0][6] if items else "Unknown"
                    display_label = f"FORM: {fid} ({assy_name})"
                    self.draw_hannan_group_row(scroll, proj, display_label, items, "RECEIVED", "#27AE60", form_id=fid)

            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not refresh data: {e}")

    def draw_hannan_header(self, parent):
        h = ctk.CTkFrame(parent, fg_color="#34495E", height=45)
        h.pack(fill="x", pady=(0, 5))
        cols = [("Group Name / Form ID", 0.05), ("Status", 0.92)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 12, "bold")).place(relx=relx, rely=0.5, anchor="w")

    def draw_hannan_group_row(self, parent, proj, label_text, items, st, color, form_id=None):
        correction_note = None
        for item in items:
            if len(item) > 13 and item[13] and str(item[13]).strip():
                correction_note = str(item[13]).strip()
                break

        if correction_note and st == "READY":
            # REKA BENTUK KHAS UNTUK CORRECTION RETURNED OLEH DC
            lines = correction_note.split('\n')
            formatted_lines = []
            for line in lines:
                if line.strip():
                    if not line.strip().startswith('-'):
                        formatted_lines.append("- " + line.strip())
                    else:
                        formatted_lines.append(line.strip())
            
            final_note = "DC Correction:\n" + "\n".join(formatted_lines)
            
            g_frame = ctk.CTkFrame(parent, fg_color="#FDEDEC", border_width=1, border_color="#E6B0AA")
            g_frame.pack(fill="x", pady=2)
            
            top_f = ctk.CTkFrame(g_frame, fg_color="transparent")
            top_f.pack(fill="x", padx=10, pady=(10, 5))
            
            ctk.CTkLabel(top_f, text=label_text, font=("Segoe UI", 12, "bold"), text_color="#2C3E50").pack(side="left", padx=5)
            ctk.CTkLabel(top_f, text=f"({len(items)} drawings)", font=("Segoe UI", 11), text_color="gray").pack(side="left", padx=5)
            
            lbl = ctk.CTkLabel(top_f, text="RETURNED", text_color="white", font=("Segoe UI", 11, "bold"), fg_color="#E74C3C", width=85, corner_radius=4)
            lbl.pack(side="right", padx=10)
            
            btn_view = ctk.CTkButton(top_f, text="VIEW LIST", width=80, height=28, font=("Segoe UI", 10, "bold"), 
                                     fg_color="#34495E", command=lambda l=label_text, i=items, fid=form_id: self.show_list_popup(l, i, fid))
            btn_view.pack(side="right", padx=5)
            
            btn_issue = ctk.CTkButton(top_f, text="RE-ISSUE", width=70, height=28, font=("Segoe UI", 10, "bold"), 
                                      fg_color="#27AE60", command=lambda p=proj, a=label_text: self.create_issue_action(p, a))
            btn_issue.pack(side="right", padx=5)
            
            bot_f = ctk.CTkFrame(g_frame, fg_color="transparent")
            bot_f.pack(fill="x", padx=15, pady=(0, 10))
            
            ctk.CTkLabel(bot_f, text=final_note, font=("Segoe UI", 11, "bold"), text_color="#C0392B", justify="left").pack(side="left")
        else:
            # PAPARAN NORMAL SEPERTI KOD ASAL ANDA
            g_frame = ctk.CTkFrame(parent, fg_color="white", height=50, border_width=1, border_color="#EAEDED")
            g_frame.pack(fill="x", pady=1)

            ctk.CTkLabel(g_frame, text=label_text, font=("Segoe UI", 12, "bold"), text_color="#2C3E50").place(relx=0.05, rely=0.5, anchor="w")
            ctk.CTkLabel(g_frame, text=f"({len(items)} drawings)", font=("Segoe UI", 11), text_color="gray").place(relx=0.45, rely=0.5, anchor="w")

            lbl = ctk.CTkLabel(g_frame, text=st, text_color="white", font=("Segoe UI", 11, "bold"), fg_color=color, width=85, corner_radius=4)
            lbl.place(relx=0.92, rely=0.5, anchor="w")

            btn_view = ctk.CTkButton(g_frame, text="VIEW LIST", width=80, height=28, font=("Segoe UI", 10, "bold"), 
                                     fg_color="#34495E", command=lambda l=label_text, i=items, fid=form_id: self.show_list_popup(l, i, fid))
            btn_view.place(relx=0.84, rely=0.5, anchor="e")

            if st == "READY":
                btn_issue = ctk.CTkButton(g_frame, text="ISSUE", width=70, height=28, font=("Segoe UI", 10, "bold"), 
                                          fg_color="#27AE60", command=lambda p=proj, a=label_text: self.create_issue_action(p, a))
                btn_issue.place(relx=0.76, rely=0.5, anchor="e")

    def show_list_popup(self, label, items, form_id=None):
        pop = ctk.CTkToplevel(self)
        pop.title(f"Drawing List: {label}")
        pop.geometry("950x550")
        pop.attributes("-topmost", True)
        
        if form_id:
            header_text = label
        else:
            header_text = f"GROUP: {label}"
            
        ctk.CTkLabel(pop, text=header_text, font=("Segoe UI", 16, "bold"), text_color="#2C3E50", justify="center").pack(pady=15)
        
        h = ctk.CTkFrame(pop, fg_color="#34495E", height=40)
        h.pack(fill="x", padx=20)
        cols = [("Drawing Name", 0.05), ("Drawing Number", 0.45), ("Rev", 0.75), ("Total", 0.85)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 12, "bold")).place(relx=relx, rely=0.5, anchor="w")

        scroll = ctk.CTkScrollableFrame(pop, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=10)

        for d in items:
            r = ctk.CTkFrame(scroll, fg_color="white", height=40, border_width=1, border_color="#F2F4F4")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=str(d[1])[:50], font=("Segoe UI", 12)).place(relx=0.05, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[2]), font=("Segoe UI", 12, "bold")).place(relx=0.45, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[3]), font=("Segoe UI", 12)).place(relx=0.75, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[4]), font=("Segoe UI", 12)).place(relx=0.85, rely=0.5, anchor="w")

    def create_issue_action(self, proj, assy_filter=None):
        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj]
            count = 0
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=2).value and not ws.cell(row=r, column=8).value:
                    if assy_filter and str(ws.cell(row=r, column=7).value) != str(assy_filter):
                        continue
                    ws.cell(row=r, column=8).value = "Hannan"
                    ws.cell(row=r, column=9).value = now
                    ws.cell(row=r, column=14).value = "" # Clear DC Correction bila re-issue
                    count += 1
            if count > 0:
                wb.save(DC_LIST_FILE)
                messagebox.showinfo("Success", f"Transmittal batch issued successfully.")
                self.refresh_hannan_data()
            else:
                messagebox.showwarning("Warning", "No new drawings to issue.")
        except PermissionError:
            messagebox.showerror("Error", "SERVER FILE BUSY: Please close Excel!")

    def setup_tasya_dashboard(self):
        top_f = ctk.CTkFrame(self.tab_trans, fg_color="transparent")
        top_f.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(top_f, text="DC Transmittal Inbox", font=("Segoe UI", 22, "bold"), text_color="#2C3E50").pack(side="left")
        ctk.CTkButton(top_f, text="REFRESH INBOX", font=("Segoe UI", 13, "bold"), command=self.refresh_tasya_inbox).pack(side="right")
        self.tasya_scroll = ctk.CTkScrollableFrame(self.tab_trans, fg_color="#F8F9F9")
        self.tasya_scroll.pack(fill="both", expand=True)
        self.refresh_tasya_inbox()

    def refresh_tasya_inbox(self):
        for w in self.tasya_scroll.winfo_children(): w.destroy()
        if not os.path.exists(DC_LIST_FILE): return
        try:
            wb = load_workbook(DC_LIST_FILE, data_only=True)
            forms = {} 
            for proj in PROJECTS:
                if proj not in wb.sheetnames: continue
                ws = wb[proj]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if r is None or not r[1]: continue
                    r_list = list(r)
                    while len(r_list) < 13: r_list.append(None)
                    if r_list[7] and not r_list[9]: # Issued but not received
                        key = (proj, r_list[6], r_list[8])
                        if key not in forms: forms[key] = []
                        forms[key].append(r_list)
            
            if not forms:
                ctk.CTkLabel(self.tasya_scroll, text="No pending transmittals found.", text_color="gray", font=("Segoe UI", 16)).pack(pady=100)
                return

            for key, items in forms.items():
                self.create_transmittal_card(key, items)
            wb.close()
        except: pass

    def create_transmittal_card(self, key, items):
        proj, assy, dt = key
        c = ctk.CTkFrame(self.tasya_scroll, border_width=1, border_color="#D5DBDB", fg_color="white")
        c.pack(fill="x", pady=8, padx=15)
        
        info = f"{proj} - {str(assy).upper()}\nIssued: {dt}  |  Total: {len(items)} Drawings"
        ctk.CTkLabel(c, text=info, font=("Segoe UI", 14, "bold"), justify="left", text_color="#2C3E50").pack(side="left", padx=25, pady=15)
        
        btn_f = ctk.CTkFrame(c, fg_color="transparent")
        btn_f.pack(side="right", padx=20)
        
        ctk.CTkButton(btn_f, text="VIEW LIST", width=120, fg_color="#34495E", font=("Segoe UI", 12, "bold"), command=lambda k=key, i=items: self.view_details_tasya(k, i)).pack(side="left", padx=5)
        
        # BUTANG CORRECTION UTK DC
        ctk.CTkButton(btn_f, text="CORRECTION", width=100, fg_color="#E67E22", font=("Segoe UI", 12, "bold"), command=lambda k=key, i=items: self.request_correction_popup(k, i)).pack(side="left", padx=5)

        ctk.CTkButton(btn_f, text="RECEIVE", width=100, fg_color="#27AE60", font=("Segoe UI", 12, "bold"), command=lambda k=key, i=items: self.receive_final_action(k, i)).pack(side="left", padx=5)

    def request_correction_popup(self, key, items):
        pop = ctk.CTkToplevel(self)
        pop.title("Request Correction")
        pop.geometry("500x350")
        pop.attributes("-topmost", True)
        
        ctk.CTkLabel(pop, text=f"Correction note for {key[0]} - {key[1]}:", font=("Segoe UI", 14, "bold")).pack(pady=15)
        
        txt_correct = ctk.CTkTextbox(pop, height=150, width=400, font=("Segoe UI", 12))
        txt_correct.pack(padx=20, pady=10)
        txt_correct.insert("1.0", "- ") # Pre-fill bullet point
        
        ctk.CTkButton(pop, text="SUBMIT CORRECTION", fg_color=COLOR_DANGER, height=45, font=("Segoe UI", 14, "bold"),
                      command=lambda: self.submit_correction(key, txt_correct.get("1.0", "end-1c").strip(), pop)).pack(pady=15)

    def submit_correction(self, key, reason, pop):
        if not reason or reason == "-":
            messagebox.showerror("Error", "Sila berikan sebab correction.")
            return
            
        proj, assy, iss_date = key
        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=9).value) == str(iss_date) and str(ws.cell(row=r, column=7).value) == str(assy):
                    ws.cell(row=r, column=8).value = ""
                    ws.cell(row=r, column=9).value = ""
                    ws.cell(row=r, column=14).value = reason # Col N untuk nota correction
            wb.save(DC_LIST_FILE)
            pop.destroy()
            messagebox.showinfo("Success", "Correction requested to Engineer.")
            self.refresh_tasya_inbox()
        except Exception as e:
            messagebox.showerror("Error", f"Server Busy: {e}")

    def view_details_tasya(self, key, items):
        proj, assy, _ = key
        pop = ctk.CTkToplevel(self); pop.title(f"Details: {proj} {assy}"); pop.geometry("1100x700"); pop.attributes("-topmost", True)
        
        h = ctk.CTkFrame(pop, fg_color="#34495E", height=45)
        h.pack(fill="x", padx=20, pady=(20, 5))
        cols = [("Drawing Name", 0.05), ("Drawing Number", 0.40), ("Rev", 0.70), ("Total", 0.78), ("Main Assy", 0.88)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 14, "bold")).place(relx=relx, rely=0.5, anchor="w")

        scroll = ctk.CTkScrollableFrame(pop, fg_color="#FDFEFE")
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        for d in items:
            row = ctk.CTkFrame(scroll, fg_color="white", height=45, border_width=1, border_color="#EAEDED")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=str(d[1])[:45], font=("Segoe UI", 13), text_color="black").place(relx=0.05, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[2]), font=("Segoe UI", 13, "bold"), text_color="#2C3E50").place(relx=0.40, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[3]), font=("Segoe UI", 13)).place(relx=0.70, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[4]), font=("Segoe UI", 13)).place(relx=0.78, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[6]), font=("Segoe UI", 13)).place(relx=0.88, rely=0.5, anchor="w")

    def receive_final_action(self, key, items):
        if not messagebox.askyesno("Confirm", f"Receive Transmittal {key[0]}?"): return
        
        proj, assy, iss_date = key
        now = datetime.now()
        f_id = f"{proj}_{assy}_{now.strftime('%d%m%Y_%H%M%S')}".upper().replace(" ", "_")
        rec_time = now.strftime("%Y-%m-%d %H:%M:%S")

        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=9).value) == str(iss_date) and str(ws.cell(row=r, column=7).value) == str(assy):
                    ws.cell(row=r, column=10).value = "Tasya"
                    ws.cell(row=r, column=11).value = rec_time
                    ws.cell(row=r, column=12).value = f_id
            wb.save(DC_LIST_FILE)
            
            self.generate_transmittal_pdf(proj, assy, items, iss_date, rec_time, f_id)
            messagebox.showinfo("Success", f"Transmittal {f_id} received and PDF saved!")
            self.refresh_tasya_inbox()
        except Exception as e:
            messagebox.showerror("Error", f"Server Busy: {e}")

    def generate_transmittal_pdf(self, proj, assy, items, iss_t, rec_t, f_id):
        folder = os.path.join(DC_ROOT, proj)
        if not os.path.exists(folder): os.makedirs(folder)
        path = os.path.join(folder, f"{f_id}.pdf")
        
        # PENGGUNAAN PDF LANDSCAPE
        pdf = TransmittalPDF(orientation='L')
        pdf.add_page()
        
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, f"FORM ID: {f_id}", ln=1)
        pdf.cell(0, 8, f"DATE: {datetime.now().strftime('%d/%m/%Y')}", ln=1)
        pdf.ln(5)
        
        # TABLE HEADER DENGAN REMARKS
        header = ["Sl", "Drawing Number", "Drawing Name", "Rev", "Total", "Approved", "Remarks"]
        data = []
        for i, d in enumerate(sorted(items, key=lambda x: str(x[2])), 1):
            appr_val = d[5]
            if hasattr(appr_val, 'strftime'): appr_str = appr_val.strftime('%Y-%m-%d')
            elif appr_val: appr_str = str(appr_val).split()[0]
            else: appr_str = ""
            
            # Ambil nilai Reason (Column M / index 12) jika ada
            rem_val = str(d[12]) if len(d) > 12 and d[12] else "-"
            data.append([str(i), str(d[2]), str(d[1])[:40], str(d[3]), str(d[4]), appr_str, rem_val[:65]])
        
        pdf.generate_table(header, data, {"proj": proj, "assy": assy})
        
        pdf.ln(10)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(95, 7, "Drawing issued by:", 0, 0)
        pdf.cell(95, 7, "Drawing received by:", 0, 1)
        pdf.set_font("Arial", "", 10)
        pdf.cell(95, 7, "Name: Hannan", 0, 0); pdf.cell(95, 7, "Name: Tasya", 0, 1)
        pdf.cell(95, 7, f"Date: {iss_t}", 0, 0); pdf.cell(95, 7, f"Date: {rec_t}", 0, 1)
        
        pdf.output(path)

if __name__ == "__main__":
    app = IntegratedApp()
    app.mainloop()