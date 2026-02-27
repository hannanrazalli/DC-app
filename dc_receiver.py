import customtkinter as ctk
import os
from openpyxl import load_workbook, Workbook
from fpdf import FPDF
from datetime import datetime, timedelta
from tkinter import messagebox

# --- CONFIGURATION PATH (UNC SERVER) ---
SERVER_BASE = r"\\192.168.2.3\Drawings"
DC_ROOT = os.path.join(SERVER_BASE, "[04] ENGINEERING TEAM", "[98] DOCUMENT CONTROL", "[05] Transmittal Form DC")
DC_LIST_FILE = os.path.join(DC_ROOT, "DC Transmittal List.xlsx")

PROJECTS = ["G10", "H10 BeraPit", "H10 TRC", "M10", "N10", "Wheel Press Machine"]

# User Credentials
USERS = {
    "Design Engineering": {"pin": "1112", "role": "ENG"},
    "Document Control": {"pin": "1234", "role": "DC"}
}

# --- PDF GENERATOR ---
class TransmittalPDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "DOCUMENT TRANSMITTAL NOTE", 0, 1, "C")
        self.ln(5)

    def generate_table(self, header, data, proj_info):
        self.set_font("Arial", "B", 8)
        content_widths = []
        for i, h_text in enumerate(header):
            w = self.get_string_width(h_text) + 6
            for row in data:
                row_w = self.get_string_width(str(row[i])) + 6
                if row_w > w: w = row_w
            content_widths.append(w)
        
        total_content_w = sum(content_widths)
        page_width = 190 
        widths = [(w / total_content_w) * page_width for w in content_widths]

        self.set_font("Arial", "B", 12)
        self.set_fill_color(144, 238, 144) 
        self.cell(0, 10, f"PROJECT: {proj_info['proj']}   |   MAIN ASSY: {proj_info['assy']}", 1, 1, 'C', 1)
        self.ln(2)

        self.set_font("Arial", "B", 8)
        self.set_fill_color(220, 220, 220)
        for i, h in enumerate(header):
            self.cell(widths[i], 10, h, 1, 0, 'C', 1)
        self.ln()
        
        self.set_font("Arial", "", 8)
        for row in data:
            if self.get_y() > 270: self.add_page()
            for i, item in enumerate(row):
                align = 'C' if i in [0, 3, 4, 5] else 'L'
                self.cell(widths[i], 7, str(item), 1, 0, align)
            self.ln()

class TransmittalApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("Light")
        self.title("LMG Engineering - Transmittal System")
        self.geometry("1250x900")
        
        self.current_user = None
        self.current_role = None
        self.project_scrolls = {} 

        self.withdraw()
        self.show_login()

    def show_login(self):
        self.login_win = ctk.CTkToplevel(self)
        self.login_win.title("Security Login")
        self.login_win.geometry("400x500")
        self.login_win.attributes("-topmost", True)
        self.login_win.protocol("WM_DELETE_WINDOW", self.quit)

        ctk.CTkLabel(self.login_win, text="SYSTEM LOGIN", font=("Segoe UI", 28, "bold"), text_color="#2C3E50").pack(pady=40)
        
        self.user_var = ctk.StringVar(value="Design Engineering")
        ctk.CTkOptionMenu(self.login_win, values=list(USERS.keys()), variable=self.user_var, height=40, font=("Segoe UI", 14)).pack(pady=10)
        
        self.pin_ent = ctk.CTkEntry(self.login_win, placeholder_text="Enter PIN Code", show="*", height=50, font=("Consolas", 24), justify="center")
        self.pin_ent.pack(pady=20, padx=60)
        self.pin_ent.focus_set()
        
        self.pin_ent.bind("<Return>", lambda e: self.verify_login())
        ctk.CTkButton(self.login_win, text="LOGIN", command=self.verify_login, height=50, font=("Segoe UI", 16, "bold"), fg_color="#2C3E50").pack(pady=30)

    def verify_login(self):
        u = self.user_var.get()
        p = self.pin_ent.get()
        if USERS.get(u)["pin"] == p:
            self.current_user = u
            self.current_role = USERS[u]["role"]
            self.login_win.destroy()
            self.deiconify()
            self.setup_ui()
        else:
            messagebox.showerror("Error", "Wrong PIN! Please try again.")

    def setup_ui(self):
        self.header_f = ctk.CTkFrame(self, height=70, fg_color="#2C3E50", corner_radius=0)
        self.header_f.pack(fill="x")
        ctk.CTkLabel(self.header_f, text="LMG ENGINEERING DC", font=("Segoe UI", 22, "bold"), text_color="white").pack(side="left", padx=30)
        ctk.CTkButton(self.header_f, text="LOGOUT", width=90, fg_color="#C0392B", font=("Segoe UI", 13, "bold"), command=self.logout).pack(side="right", padx=20)

        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True, padx=20, pady=20)

        if self.current_role == "ENG":
            self.setup_hannan_dashboard()
        else:
            self.setup_tasya_dashboard()

    def logout(self):
        if messagebox.askyesno("Logout", "Do you want to exit?"):
            self.current_user = None
            for w in self.winfo_children(): w.destroy()
            self.withdraw()
            self.show_login()

    def setup_hannan_dashboard(self):
        top_bar = ctk.CTkFrame(self.main_container, fg_color="transparent")
        top_bar.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(top_bar, text="Engineering Transmittal Dashboard", font=("Segoe UI", 20, "bold"), text_color="#2C3E50").pack(side="left")
        ctk.CTkButton(top_bar, text="REFRESH ALL", font=("Segoe UI", 13, "bold"), command=self.refresh_hannan_data).pack(side="right")

        self.tabview = ctk.CTkTabview(self.main_container, corner_radius=15, segmented_button_selected_color="#2C3E50")
        self.tabview.pack(fill="both", expand=True)
        
        self.project_scrolls = {}
        for proj in PROJECTS:
            tab = self.tabview.add(proj)
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(1, weight=1)
            
            info_f = ctk.CTkFrame(tab, fg_color="transparent")
            info_f.grid(row=0, column=0, sticky="ew", pady=10, padx=15)
            
            self.project_scrolls[proj] = {
                "scroll": ctk.CTkScrollableFrame(tab, fg_color="#FBFCFC"),
                "note_label": ctk.CTkLabel(info_f, text="", font=("Segoe UI", 12, "bold")),
                "issue_btn": ctk.CTkButton(info_f, text="Create & Issue (All)", fg_color="#27AE60", font=("Segoe UI", 13, "bold"), command=lambda p=proj: self.create_issue_action(p))
            }
            self.project_scrolls[proj]["note_label"].pack(side="left")
            self.project_scrolls[proj]["issue_btn"].pack(side="right")
            self.project_scrolls[proj]["scroll"].grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

        self.refresh_hannan_data()

    def refresh_hannan_data(self):
        if not os.path.exists(DC_LIST_FILE): return
        try:
            wb = load_workbook(DC_LIST_FILE, data_only=True)
            now_dt = datetime.now()
            
            for proj in PROJECTS:
                if proj not in self.project_scrolls: continue
                scroll = self.project_scrolls[proj]["scroll"]
                note = self.project_scrolls[proj]["note_label"]
                
                for w in scroll.winfo_children(): w.destroy()
                
                if proj not in wb.sheetnames: continue
                ws = wb[proj]
                
                raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
                rows = []
                for r in raw_rows:
                    if r is None or not r[1]: continue
                    r_list = list(r)
                    while len(r_list) < 13: r_list.append(None)
                    rows.append(r_list)
                
                # --- Grouping logic for all states ---
                unissued_groups = {} # Belum Isu: Group by Main Assy
                pending_groups = {}  # ISSUED: Group by (Main Assy + Issue Time)
                received_groups = {} # RECEIVED: Group by Form ID (Unique per transmittal)

                for r in rows:
                    assy = str(r[6])
                    iss_time = str(r[8]) # Col I
                    f_id = str(r[11])    # Col L
                    
                    if not r[7]: # BELUM ISU
                        if assy not in unissued_groups: unissued_groups[assy] = []
                        unissued_groups[assy].append(r)
                    elif not r[9]: # ISSUED but not Received
                        key = (assy, iss_time)
                        if key not in pending_groups: pending_groups[key] = []
                        pending_groups[key].append(r)
                    else: # RECEIVED
                        rec_time_val = r[10]
                        try:
                            if isinstance(rec_time_val, str): rec_dt = datetime.strptime(rec_time_val, "%Y-%m-%d %H:%M:%S")
                            else: rec_dt = rec_time_val
                            if rec_dt and (now_dt - rec_dt) < timedelta(hours=1):
                                if f_id not in received_groups: received_groups[f_id] = []
                                received_groups[f_id].append(r)
                        except: pass

                # Update Status Counter
                unissued_count = len(unissued_groups.keys())
                pending_count = len(pending_groups.keys())
                note.configure(text=f"Unissued Groups: {unissued_count}  |  Pending DC: {pending_count} Batches", 
                               text_color="#C0392B" if unissued_count > 0 else "#27AE60")
                
                # Header
                self.draw_hannan_header(scroll)

                # 1. BELUM ISU (Groups by Assy)
                for assy, items in unissued_groups.items():
                    self.draw_hannan_group_row(scroll, proj, assy, items, "READY", "#3498DB")

                # 2. PENDING RECEIVED (Groups by Assy + Time)
                for (assy, time), items in pending_groups.items():
                    display_label = f"{assy} (Issued: {time})"
                    self.draw_hannan_group_row(scroll, proj, display_label, items, "ISSUED", "#E67E22")

                # 3. RECEIVED (Groups by Form ID)
                for fid, items in received_groups.items():
                    assy_name = items[0][6] if items else "Unknown"
                    display_label = f"FORM: {fid} ({assy_name})"
                    self.draw_hannan_group_row(scroll, proj, display_label, items, "RECEIVED", "#27AE60", form_id=fid)

            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not refresh data: {e}")

    def draw_hannan_header(self, parent):
        h = ctk.CTkFrame(parent, fg_color="#34495E", height=45)
        h.pack(fill="x", pady=(0, 5))
        cols = [("Group Name / Form ID", 0.05), ("Status", 0.92)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 12, "bold")).place(relx=relx, rely=0.5, anchor="w")

    def draw_hannan_group_row(self, parent, proj, label_text, items, st, color, form_id=None):
        """Draws a grouped row with a pop-up View List."""
        g_frame = ctk.CTkFrame(parent, fg_color="white", height=50, border_width=1, border_color="#EAEDED")
        g_frame.pack(fill="x", pady=1)

        ctk.CTkLabel(g_frame, text=label_text, font=("Segoe UI", 12, "bold"), text_color="#2C3E50").place(relx=0.05, rely=0.5, anchor="w")
        ctk.CTkLabel(g_frame, text=f"({len(items)} drawings)", font=("Segoe UI", 11), text_color="gray").place(relx=0.45, rely=0.5, anchor="w")

        lbl = ctk.CTkLabel(g_frame, text=st, text_color="white", font=("Segoe UI", 11, "bold"), fg_color=color, width=85, corner_radius=4)
        lbl.place(relx=0.92, rely=0.5, anchor="w")

        btn_view = ctk.CTkButton(g_frame, text="VIEW LIST", width=80, height=28, font=("Segoe UI", 10, "bold"), 
                                 fg_color="#34495E", command=lambda l=label_text, i=items, fid=form_id: self.show_list_popup(l, i, fid))
        btn_view.place(relx=0.84, rely=0.5, anchor="e")

        if st == "READY":
            btn_issue = ctk.CTkButton(g_frame, text="ISSUE", width=70, height=28, font=("Segoe UI", 10, "bold"), 
                                      fg_color="#27AE60", command=lambda p=proj, a=label_text: self.create_issue_action(p, a))
            btn_issue.place(relx=0.76, rely=0.5, anchor="e")

    def show_list_popup(self, label, items, form_id=None):
        """Shows a Toplevel popup with the list of drawings for the group."""
        pop = ctk.CTkToplevel(self)
        pop.title(f"Drawing List: {label}")
        pop.geometry("950x550")
        pop.attributes("-topmost", True)
        
        # Penyesuaian: Gunakan label asal sahaja jika form_id wujud untuk elakkan ID dipaparkan dua kali
        if form_id:
            header_text = label
        else:
            header_text = f"GROUP: {label}"
            
        ctk.CTkLabel(pop, text=header_text, font=("Segoe UI", 16, "bold"), text_color="#2C3E50", justify="center").pack(pady=15)
        
        h = ctk.CTkFrame(pop, fg_color="#34495E", height=40)
        h.pack(fill="x", padx=20)
        cols = [("Drawing Name", 0.05), ("Drawing Number", 0.45), ("Rev", 0.75), ("Total", 0.85)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 12, "bold")).place(relx=relx, rely=0.5, anchor="w")

        scroll = ctk.CTkScrollableFrame(pop, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=10)

        for d in items:
            r = ctk.CTkFrame(scroll, fg_color="white", height=40, border_width=1, border_color="#F2F4F4")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=str(d[1])[:50], font=("Segoe UI", 12)).place(relx=0.05, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[2]), font=("Segoe UI", 12, "bold")).place(relx=0.45, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[3]), font=("Segoe UI", 12)).place(relx=0.75, rely=0.5, anchor="w")
            ctk.CTkLabel(r, text=str(d[4]), font=("Segoe UI", 12)).place(relx=0.85, rely=0.5, anchor="w")

    def create_issue_action(self, proj, assy_filter=None):
        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj]
            count = 0
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=2).value and not ws.cell(row=r, column=8).value:
                    if assy_filter and str(ws.cell(row=r, column=7).value) != str(assy_filter):
                        continue
                    ws.cell(row=r, column=8).value = "Hannan"
                    ws.cell(row=r, column=9).value = now
                    count += 1
            if count > 0:
                wb.save(DC_LIST_FILE)
                messagebox.showinfo("Success", f"Transmittal batch issued successfully.")
                self.refresh_hannan_data()
            else:
                messagebox.showwarning("Warning", "No new drawings to issue.")
        except PermissionError:
            messagebox.showerror("Error", "SERVER FILE BUSY: Please close Excel!")

    def setup_tasya_dashboard(self):
        top_f = ctk.CTkFrame(self.main_container, fg_color="transparent")
        top_f.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(top_f, text="DC Transmittal Inbox", font=("Segoe UI", 22, "bold"), text_color="#2C3E50").pack(side="left")
        ctk.CTkButton(top_f, text="REFRESH INBOX", font=("Segoe UI", 13, "bold"), command=self.refresh_tasya_inbox).pack(side="right")
        self.tasya_scroll = ctk.CTkScrollableFrame(self.main_container, fg_color="#F8F9F9")
        self.tasya_scroll.pack(fill="both", expand=True)
        self.refresh_tasya_inbox()

    def refresh_tasya_inbox(self):
        for w in self.tasya_scroll.winfo_children(): w.destroy()
        if not os.path.exists(DC_LIST_FILE): return
        try:
            wb = load_workbook(DC_LIST_FILE, data_only=True)
            forms = {}
            for proj in PROJECTS:
                if proj not in wb.sheetnames: continue
                ws = wb[proj]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if r is None or not r[1]: continue
                    r_list = list(r)
                    while len(r_list) < 13: r_list.append(None)
                    if r_list[7] and not r_list[9]: # Issued but not received
                        key = (proj, r_list[6], r_list[8])
                        if key not in forms: forms[key] = []
                        forms[key].append(r_list)
            if not forms:
                ctk.CTkLabel(self.tasya_scroll, text="No pending transmittals found.", text_color="gray", font=("Segoe UI", 16)).pack(pady=100)
                return
            for key, items in forms.items():
                self.create_transmittal_card(key, items)
            wb.close()
        except: pass

    def create_transmittal_card(self, key, items):
        proj, assy, dt = key
        c = ctk.CTkFrame(self.tasya_scroll, border_width=1, border_color="#D5DBDB", fg_color="white")
        c.pack(fill="x", pady=8, padx=15)
        info = f"{proj} - {str(assy).upper()}\nIssued: {dt}  |  Total: {len(items)} Drawings"
        ctk.CTkLabel(c, text=info, font=("Segoe UI", 14, "bold"), justify="left", text_color="#2C3E50").pack(side="left", padx=25, pady=15)
        btn_f = ctk.CTkFrame(c, fg_color="transparent")
        btn_f.pack(side="right", padx=20)
        ctk.CTkButton(btn_f, text="VIEW LIST", width=120, fg_color="#34495E", font=("Segoe UI", 12, "bold"), command=lambda k=key, i=items: self.view_details_tasya(k, i)).pack(side="left", padx=5)
        ctk.CTkButton(btn_f, text="RECEIVE", width=100, fg_color="#27AE60", font=("Segoe UI", 12, "bold"), command=lambda k=key, i=items: self.receive_final_action(k, i)).pack(side="left", padx=5)

    def view_details_tasya(self, key, items):
        proj, assy, _ = key
        pop = ctk.CTkToplevel(self); pop.title(f"Details: {proj} {assy}"); pop.geometry("1100x700"); pop.attributes("-topmost", True)
        h = ctk.CTkFrame(pop, fg_color="#34495E", height=45)
        h.pack(fill="x", padx=20, pady=(20, 5))
        cols = [("Drawing Name", 0.05), ("Drawing Number", 0.40), ("Rev", 0.70), ("Total", 0.78), ("Main Assy", 0.88)]
        for text, relx in cols:
            ctk.CTkLabel(h, text=text, text_color="white", font=("Segoe UI", 14, "bold")).place(relx=relx, rely=0.5, anchor="w")
        scroll = ctk.CTkScrollableFrame(pop, fg_color="#FDFEFE")
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        for d in items:
            row = ctk.CTkFrame(scroll, fg_color="white", height=45, border_width=1, border_color="#EAEDED")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=str(d[1])[:45], font=("Segoe UI", 13), text_color="black").place(relx=0.05, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[2]), font=("Segoe UI", 13, "bold"), text_color="#2C3E50").place(relx=0.40, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[3]), font=("Segoe UI", 13)).place(relx=0.70, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[4]), font=("Segoe UI", 13)).place(relx=0.78, rely=0.5, anchor="w")
            ctk.CTkLabel(row, text=str(d[6]), font=("Segoe UI", 13)).place(relx=0.88, rely=0.5, anchor="w")

    def receive_final_action(self, key, items):
        if not messagebox.askyesno("Confirm", f"Receive Transmittal {key[0]}?"): return
        proj, assy, iss_date = key
        now = datetime.now()
        f_id = f"{proj}_{assy}_{now.strftime('%d%m%Y_%H%M%S')}".upper().replace(" ", "_")
        rec_time = now.strftime("%Y-%m-%d %H:%M:%S")
        try:
            wb = load_workbook(DC_LIST_FILE)
            ws = wb[proj]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=9).value) == str(iss_date) and ws.cell(row=r, column=7).value == assy:
                    ws.cell(row=r, column=10).value = "Tasya"
                    ws.cell(row=r, column=11).value = rec_time
                    ws.cell(row=r, column=12).value = f_id
            wb.save(DC_LIST_FILE)
            self.generate_transmittal_pdf(proj, assy, items, iss_date, rec_time, f_id)
            messagebox.showinfo("Success", f"Form {f_id} saved successfully!")
            self.refresh_tasya_inbox()
        except: messagebox.showerror("Error", "Server Busy!")

    def generate_transmittal_pdf(self, proj, assy, items, iss_t, rec_t, f_id):
        folder = os.path.join(DC_ROOT, proj)
        if not os.path.exists(folder): os.makedirs(folder)
        path = os.path.join(folder, f"{f_id}.pdf")
        pdf = TransmittalPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, f"FORM ID: {f_id}", ln=1)
        pdf.cell(0, 8, f"DATE: {datetime.now().strftime('%d/%m/%Y')}", ln=1)
        pdf.ln(5)
        header = ["Sl", "Drawing Number", "Drawing Name", "Rev", "Total", "Approved"]
        data = []
        for i, d in enumerate(sorted(items, key=lambda x: str(x[2])), 1):
            appr_val = d[5]
            if hasattr(appr_val, 'strftime'): appr_str = appr_val.strftime('%Y-%m-%d')
            elif appr_val: appr_str = str(appr_val).split()[0]
            else: appr_str = ""
            data.append([str(i), str(d[2]), str(d[1])[:40], str(d[3]), str(d[4]), appr_str])
        
        pdf.generate_table(header, data, {"proj": proj, "assy": assy})
        pdf.ln(10)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(95, 7, "Drawing issued by:", 0, 0)
        pdf.cell(95, 7, "Drawing received by:", 0, 1)
        pdf.set_font("Arial", "", 10)
        pdf.cell(95, 7, f"Name: Hannan", 0, 0); pdf.cell(95, 7, f"Name: Tasya", 0, 1)
        pdf.cell(95, 7, f"Date: {iss_t}", 0, 0); pdf.cell(95, 7, f"Date: {rec_t}", 0, 1)
        pdf.output(path)

if __name__ == "__main__":
    app = TransmittalApp()
    app.mainloop()