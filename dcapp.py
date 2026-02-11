import customtkinter as ctk
import psycopg2
import os
import threading
import shutil  # Dari Kod 2
import time    # Dari Kod 2
from tkinter import messagebox
from datetime import date, datetime
from tkcalendar import DateEntry
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from PIL import Image, ImageTk 

# --- APPEARANCE SETUP ---
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("dark-blue")

# --- CONFIGURATION (Kombinasi Kod 1 & Kod 2) ---
BASE_PATH = r"C:\Users\HP\Documents\[01] Document Control"
DB_CONFIG = {"dbname": "dcde", "user": "postgres", "password": "1234", "host": "localhost"}
MASTER_FILE = "Document Control.xlsx"
LOGO_FILENAME = "LMG Locomotive Logo.jpeg" 

# CONFIGURATION DARI KOD 2
SERVER_PATH = r"Y:\[04] ENGINEERING TEAM\[98] DOCUMENT CONTROL"
FILE_MAPPING = {
    "Document Control.xlsx": "All Projects Drawings Data.xlsx",
    "H10 Berapit.xlsx": "H10 BeraPit Drawing List.xlsx",
    "H10 TRC.xlsx": "H10 TRC Drawing List.xlsx",
    "M10.xlsx": "M10(N) Drawing List.xlsx",
    "N10.xlsx": "N10(N) Drawing List.xlsx",
    "Wheel Press Machine.xlsx": "Wheel Press Machine Drawing List.xlsx",
    "G10.xlsx": "G10 Drawing List.xlsx"
}

# --- PDF DRAWING CONFIGURATION ---
# Root folder untuk semua drawings
BASE_DRAWINGS_PATH = r"Y:\[04] ENGINEERING TEAM\[98] DOCUMENT CONTROL\[00] 2D Drawings - Signed Copy"

# PROJECT CONFIGURATION
PROJECTS = ["H10 TRC", "H10 BeraPit", "M10", "N10", "G10", "Wheel Press Machine"]

PROJ_MAP = {
    "H10 TRC": "H10", 
    "H10 BeraPit": "H10", 
    "M10": "M10", 
    "N10": "N10", 
    "G10": "G10 Drawing List", 
    "Wheel Press Machine": "WPM"
}

REVERSE_PROJ_MAP = {}
for k, v in PROJ_MAP.items():
    if v not in REVERSE_PROJ_MAP: 
        REVERSE_PROJ_MAP[v] = k

ASSEMBLIES = ["Bogie", "Underframe", "Cabin", "Engine Hood", "Radiator Hood", "Muffler", "Gear Case", "Water Tank", "Battery Box", "Fuel Tank", "Sandbox"]
ENGINEER_LIST = ["Baskaran", "Sathish", "Harrison", "Hannan", "Gokul", "Vimal", "Ram", "Vishwa", "Bruno"]
REMARKS_LIST = ["New", "Revised", "Re-Release", "-"]
# BATCH_LIST dari kod anda
BATCH_LIST = ["-", "1", "2", "N", "R"]
MASTER_HEADERS = ["Project", "Country", "Batch", "Main Assembly", "Drawing Name", "Part Number", "Revision", "Total Sheets", "Engineer", "Date Approved", "Remarks"]

# --- COLORS & FONTS ---
COLOR_PRIMARY = "#2C3E50"    
COLOR_ACCENT = "#3498DB"     
COLOR_SUCCESS = "#27AE60"    
COLOR_DANGER = "#C0392B"     
COLOR_WARNING = "#F39C12"    
COLOR_INFO = "#1ABC9C"       
COLOR_BG = "#ECF0F1"         
COLOR_CARD = "#FFFFFF"       
COLOR_TEXT = "#34495E"       

FONT_HEADER = ("Segoe UI", 24, "bold")
FONT_SECTION = ("Segoe UI", 16, "bold")
FONT_LABEL = ("Segoe UI", 12, "bold")
FONT_INPUT = ("Segoe UI", 13)
FONT_FEEDBACK = ("Segoe UI", 16, "bold") 
FONT_ERROR = ("Segoe UI", 10, "bold")

CREAM_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

class DCDEApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Design Engineering Data Entry")
        self.geometry("1024x920") 
        self.minsize(950, 800) 
        self.configure(fg_color=COLOR_BG)
        
        try:
            if os.path.exists(LOGO_FILENAME):
                self.window_icon = ImageTk.PhotoImage(file=LOGO_FILENAME)
                self.wm_iconphoto(False, self.window_icon)
        except Exception as e:
            print(f"Warning: Could not set window icon. {e}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) 
        
        self.setup_ui()

    def setup_ui(self):
        # Header
        self.header_frame = ctk.CTkFrame(self, fg_color=COLOR_PRIMARY, corner_radius=0, height=85)
        self.header_frame.grid(row=0, column=0, sticky="ew")
        self.header_frame.grid_propagate(False)

        self.title_label = ctk.CTkLabel(self.header_frame, text="ENGINEERING DOCUMENT CONTROL", font=FONT_HEADER, text_color="white")
        self.title_label.pack(side="left", padx=25, pady=15)
        
        self.subtitle_label = ctk.CTkLabel(self.header_frame, text="|  Data Entry System", font=("Segoe UI", 14), text_color="#BDC3C7")
        self.subtitle_label.pack(side="left", pady=20)

        # Right Side Container
        self.right_side_container = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        self.right_side_container.pack(side="right", padx=30)

        self.publish_ui_frame = ctk.CTkFrame(self.right_side_container, fg_color="transparent")
        self.publish_ui_frame.pack(side="left", padx=(0, 20))

        self.btn_publish = ctk.CTkButton(self.publish_ui_frame, text="PUBLISH TO SERVER", 
                                       fg_color=COLOR_INFO, hover_color="#16A085", 
                                       font=("Segoe UI", 12, "bold"), height=32,
                                       command=self.publish_to_server_thread)
        self.btn_publish.pack(pady=(0, 2))

        self.lbl_last_sync = ctk.CTkLabel(self.publish_ui_frame, text="Last Sync: Never", 
                                        font=("Segoe UI", 10), text_color="#BDC3C7")
        self.lbl_last_sync.pack()

        try:
            if os.path.exists(LOGO_FILENAME):
                pil_img = Image.open(LOGO_FILENAME)
                target_height = 40
                aspect_ratio = pil_img.width / pil_img.height
                target_width = int(target_height * aspect_ratio)
                self.header_logo = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(target_width, target_height))
                self.logo_label = ctk.CTkLabel(self.right_side_container, text="", image=self.header_logo)
                self.logo_label.pack(side="right")
        except Exception as e:
            print(f"Error loading logo: {e}")

        # --- TAB VIEW SYSTEM ---
        self.tabview = ctk.CTkTabview(self, fg_color=COLOR_CARD, border_width=1, border_color="#BDC3C7", corner_radius=15)
        self.tabview.grid(row=1, column=0, sticky="nsew", padx=20, pady=(10, 20))
        
        self.tab_entry = self.tabview.add("DATA ENTRY")
        self.tab_drawings = self.tabview.add("VIEW DRAWINGS")

        # --- TAB 1: DATA ENTRY ---
        self.tab_entry.columnconfigure((0, 1), weight=1)
        
        self.create_section_header("1. PROJECT DETAILS", row=0, parent=self.tab_entry)
        self.add_input_field(label="Project Name", row=1, col=0, parent=self.tab_entry)
        self.proj_v = ctk.StringVar(value=PROJECTS[0])
        self.proj_drop = ctk.CTkOptionMenu(self.tab_entry, values=PROJECTS, variable=self.proj_v, command=self.update_logic, 
                                           fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.proj_drop.grid(row=2, column=0, padx=20, pady=(5, 15), sticky="ew")

        self.add_input_field(label="Batch Code", row=1, col=1, parent=self.tab_entry)
        self.batch_v = ctk.StringVar(value="-")
        self.batch_drop = ctk.CTkOptionMenu(self.tab_entry, values=BATCH_LIST, variable=self.batch_v,
                                            fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.batch_drop.grid(row=2, column=1, padx=20, pady=(5, 15), sticky="ew")

        self.create_section_header("2. DRAWING & TECHNICAL INFORMATION", row=3, parent=self.tab_entry)
        self.add_input_field(label="Main Assembly", row=4, col=0, parent=self.tab_entry)
        self.assembly_v = ctk.StringVar(value=ASSEMBLIES[0])
        self.assembly_drop = ctk.CTkOptionMenu(self.tab_entry, values=ASSEMBLIES, variable=self.assembly_v, 
                                                fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.assembly_drop.grid(row=5, column=0, padx=20, pady=(5, 10), sticky="ew")

        self.add_input_field(label="Drawing Name", row=4, col=1, parent=self.tab_entry)
        self.draw_ent = ctk.CTkEntry(self.tab_entry, font=FONT_INPUT, height=32, placeholder_text="Full drawing title")
        self.draw_ent.grid(row=5, column=1, padx=20, pady=(5, 0), sticky="ew")
        self.draw_ent.bind("<KeyRelease>", lambda e: [self.to_uppercase(e, self.draw_ent), self.clear_field_error("draw")])
        self.err_draw = ctk.CTkLabel(self.tab_entry, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_draw.grid(row=6, column=1, padx=25, sticky="w")

        self.add_input_field(label="Part Number", row=7, col=0, parent=self.tab_entry)
        self.part_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.part_frame.grid(row=8, column=0, padx=20, pady=(5, 0), sticky="ew")
        self.part_ent = ctk.CTkEntry(self.part_frame, font=FONT_INPUT, height=32, placeholder_text="e.g. H10-100-001")
        self.part_ent.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.part_ent.bind("<KeyRelease>", lambda e: [self.on_part_input(e), self.clear_field_error("part")])
        self.btn_check_part = ctk.CTkButton(self.part_frame, text="Search", width=60, height=32, 
                                            fg_color="#7f8c8d", hover_color="#95a5a6", 
                                            command=self.check_part_existence_thread)
        self.btn_check_part.pack(side="right")
        self.err_part = ctk.CTkLabel(self.tab_entry, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_part.grid(row=9, column=0, padx=25, sticky="w")

        self.sub_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.sub_frame.grid(row=7, column=1, rowspan=3, padx=20, pady=0, sticky="ew")
        self.sub_frame.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkLabel(self.sub_frame, text="Revision", font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=0, column=0, sticky="w")
        self.rev_ent = ctk.CTkEntry(self.sub_frame, font=FONT_INPUT, height=32)
        self.rev_ent.grid(row=1, column=0, padx=(0, 10), pady=(5, 0), sticky="ew")
        self.rev_ent.bind("<KeyRelease>", lambda e: [self.auto_remark_logic(e), self.clear_field_error("rev")])
        self.err_rev = ctk.CTkLabel(self.sub_frame, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_rev.grid(row=2, column=0, padx=(0, 10), sticky="w")

        ctk.CTkLabel(self.sub_frame, text="Total Sheets", font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=0, column=1, sticky="w")
        self.total_ent = ctk.CTkEntry(self.sub_frame, font=FONT_INPUT, height=32)
        self.total_ent.grid(row=1, column=1, padx=(10, 0), pady=(5, 0), sticky="ew")
        self.total_ent.bind("<KeyRelease>", lambda e: self.clear_field_error("total"))
        self.err_total = ctk.CTkLabel(self.sub_frame, text="", font=FONT_ERROR, text_color=COLOR_DANGER)
        self.err_total.grid(row=2, column=1, padx=(10, 0), sticky="w")

        self.create_section_header("3. APPROVAL & STATUS", row=10, parent=self.tab_entry)
        self.add_input_field(label="Responsible Engineer", row=11, col=0, parent=self.tab_entry)
        self.eng_v = ctk.StringVar(value=ENGINEER_LIST[0])
        self.eng_drop = ctk.CTkOptionMenu(self.tab_entry, values=ENGINEER_LIST, variable=self.eng_v, 
                                          fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.eng_drop.grid(row=12, column=0, padx=20, pady=(5, 10), sticky="ew")

        self.add_input_field(label="Date Approved", row=11, col=1, parent=self.tab_entry)
        self.date_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.date_frame.grid(row=12, column=1, padx=20, pady=(5, 10), sticky="w")
        self.date_picker = DateEntry(self.date_frame, width=20, background=COLOR_PRIMARY, foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', font=("Arial", 11))
        self.date_picker.pack(side="left", padx=(0, 15), ipady=3)
        ctk.CTkButton(self.date_frame, text="Set Today", width=100, height=30, fg_color="#95a5a6", hover_color="#7f8c8d", command=self.set_today).pack(side="left")

        self.add_input_field(label="Remarks", row=13, col=0, parent=self.tab_entry)
        self.remark_v = ctk.StringVar(value="New")
        self.remark_drop = ctk.CTkOptionMenu(self.tab_entry, values=REMARKS_LIST, variable=self.remark_v, 
                                             fg_color=COLOR_PRIMARY, button_color=COLOR_ACCENT, font=FONT_INPUT, height=32)
        self.remark_drop.grid(row=14, column=0, padx=20, pady=(5, 15), sticky="ew")

        # --- HIGHLIGHT / FEEDBACK AREA ---
        self.feedback_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent", height=50)
        self.feedback_frame.grid(row=15, column=0, columnspan=2, pady=(10, 5), padx=20, sticky="ew")
        self.lbl_feedback = ctk.CTkLabel(self.feedback_frame, text="", font=FONT_FEEDBACK, text_color="white", corner_radius=6)
        self.lbl_feedback.pack(fill="both", ipady=10)

        self.btn_frame = ctk.CTkFrame(self.tab_entry, fg_color="transparent")
        self.btn_frame.grid(row=16, column=0, columnspan=2, pady=(10, 20), padx=20, sticky="ew")
        self.btn_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        ctk.CTkButton(self.btn_frame, text="SUBMIT DATA", fg_color=COLOR_SUCCESS, hover_color="#219150", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.submit).grid(row=0, column=0, padx=(0, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="CLEAR FORM", fg_color=COLOR_DANGER, hover_color="#A93226", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.clear_all).grid(row=0, column=1, padx=(5, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="OPEN FOLDER", fg_color=COLOR_WARNING, hover_color="#D35400", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.open_folder).grid(row=0, column=2, padx=(5, 5), sticky="ew")
        ctk.CTkButton(self.btn_frame, text="OPEN EXCEL", fg_color=COLOR_INFO, hover_color="#16A085", height=50, corner_radius=8, font=("Segoe UI", 15, "bold"), command=self.open_project_file).grid(row=0, column=3, padx=(5, 0), sticky="ew")

        # --- TAB 2: VIEW DRAWINGS ---
        self.tab_drawings.columnconfigure(0, weight=1)
        self.create_section_header("VIEW SIGNED DRAWINGS (SERVER)", row=0, parent=self.tab_drawings)
        
        self.search_draw_frame = ctk.CTkFrame(self.tab_drawings, fg_color="transparent")
        self.search_draw_frame.grid(row=1, column=0, padx=40, pady=40, sticky="ew")
        self.search_draw_frame.columnconfigure(0, weight=1)

        ctk.CTkLabel(self.search_draw_frame, text="Enter Part Number to Search Signed Copy:", font=FONT_LABEL).grid(row=0, column=0, sticky="w", pady=(0, 10))
        self.pdf_search_ent = ctk.CTkEntry(self.search_draw_frame, font=("Segoe UI", 16), height=45, placeholder_text="e.g. H10-100-001")
        self.pdf_search_ent.grid(row=1, column=0, sticky="ew", padx=(0, 15))
        self.pdf_search_ent.bind("<KeyRelease>", lambda e: self.to_uppercase(e, self.pdf_search_ent))

        self.btn_search_pdf = ctk.CTkButton(self.search_draw_frame, text="SEARCH DRAWING", fg_color=COLOR_ACCENT, 
                                           height=45, width=200, font=("Segoe UI", 13, "bold"), command=self.search_pdf_thread)
        self.btn_search_pdf.grid(row=1, column=1)

        self.pdf_status_lbl = ctk.CTkLabel(self.tab_drawings, text="", font=("Segoe UI", 12))
        self.pdf_status_lbl.grid(row=2, column=0, pady=10)

        # Status Bar
        self.status_bar = ctk.CTkFrame(self, height=30, fg_color="#BDC3C7", corner_radius=0)
        self.status_bar.grid(row=2, column=0, sticky="ew")
        self.status_label = ctk.CTkLabel(self.status_bar, text="System Ready", font=("Consolas", 11), text_color="#2C3E50")
        self.status_label.pack(side="left", padx=20)

    # --- CLASS METHODS ---

    def create_section_header(self, text, row, parent):
        label = ctk.CTkLabel(parent, text=text, font=FONT_SECTION, text_color=COLOR_PRIMARY, anchor="w")
        label.grid(row=row, column=0, columnspan=2, padx=20, pady=(15, 5), sticky="w")
        line = ctk.CTkFrame(parent, height=2, fg_color="#E0E0E0")
        line.grid(row=row+1, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="ew")

    def add_input_field(self, label, row, col, parent):
        ctk.CTkLabel(parent, text=label, font=FONT_LABEL, text_color=COLOR_TEXT).grid(row=row, column=col, padx=20, pady=(5, 0), sticky="w")

    def clear_field_error(self, field):
        if field == "draw": self.err_draw.configure(text="")
        elif field == "part": self.err_part.configure(text="")
        elif field == "rev": self.err_rev.configure(text="")
        elif field == "total": self.err_total.configure(text="")

    def update_status(self, m, c=COLOR_TEXT): 
        self.after(0, lambda: self.status_label.configure(text=f"STATUS: {m}", text_color=c))

    def trigger_feedback(self, t="success", m="Saved Successfully"):
        c = COLOR_SUCCESS if t == "success" else COLOR_DANGER if t == "error" else COLOR_INFO
        self.lbl_feedback.configure(text=m, fg_color=c)
        self.after(5000, lambda: self.lbl_feedback.configure(text="", fg_color="transparent"))

    def set_process_highlight(self, message, color=COLOR_INFO):
        self.lbl_feedback.configure(text=message, fg_color=color)

    def open_folder(self):
        if os.path.exists(BASE_PATH): os.startfile(BASE_PATH)
        else: messagebox.showerror("Error", "Folder not found")

    def open_project_file(self):
        fp = os.path.join(BASE_PATH, f"{self.proj_v.get()}.xlsx")
        if os.path.exists(fp): os.startfile(fp)
        else: messagebox.showerror("Error", "File not found")

    def to_uppercase(self, event, widget):
        if event.keysym in ["Control_L", "Control_R", "Shift_L", "Shift_R", "Left", "Right"]: return
        pos = widget.index(ctk.INSERT); val = widget.get().upper()
        if widget.get() != val: widget.delete(0, 'end'); widget.insert(0, val); widget.icursor(pos)

    def on_part_input(self, event):
        self.to_uppercase(event, self.part_ent)
        if self.proj_v.get() != "Wheel Press Machine" and "(N)" in self.part_ent.get(): self.batch_v.set("N")

    def auto_remark_logic(self, event):
        val = self.rev_ent.get()
        if val.isdigit(): self.remark_v.set("Revised" if int(val) >= 1 else "New")

    def update_logic(self, choice):
        self.batch_drop.configure(state="normal")
        if choice == "Wheel Press Machine":
            self.batch_v.set("-"); self.batch_drop.configure(state="disabled") 
            self.assembly_drop.configure(values=["Wheel Press"]); self.assembly_v.set("Wheel Press")
            self.eng_v.set("Baskaran"); self.eng_drop.configure(values=["Baskaran"])
        elif choice in ["M10", "N10"]:
            self.batch_v.set("-"); self.assembly_drop.configure(values=ASSEMBLIES + ["Auxiliary Hood"])
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)
        elif choice == "H10 TRC":
            self.batch_v.set("2"); self.assembly_drop.configure(values=ASSEMBLIES)
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)
        else:
            self.assembly_drop.configure(values=ASSEMBLIES)
            if self.assembly_v.get() == "Wheel Press": self.assembly_v.set(ASSEMBLIES[0])
            self.eng_drop.configure(values=ENGINEER_LIST)

    def set_today(self): self.date_picker.set_date(date.today())

    def clear_all(self):
        self.update_logic(self.proj_v.get())
        for w in [self.draw_ent, self.part_ent, self.rev_ent, self.total_ent]: w.delete(0, 'end')
        self.err_draw.configure(text=""); self.err_part.configure(text=""); self.err_rev.configure(text=""); self.err_total.configure(text="")
        self.lbl_feedback.configure(text="", fg_color="transparent")

    # --- SEARCH LOGIC ---
    def check_part_existence_thread(self):
        self.btn_check_part.configure(state="disabled", text="...")
        self.set_process_highlight("🔍 SEARCHING DATABASE & FILES... PLEASE WAIT", COLOR_WARNING)
        threading.Thread(target=self._run_global_search, args=(self.part_ent.get().strip().upper(),), daemon=True).start()

    def _run_global_search(self, part):
        if not part: self.after(0, lambda: self._post_check([])); return
        matches = []
        m_path = os.path.join(BASE_PATH, MASTER_FILE)
        try:
            if os.path.exists(m_path):
                wb = load_workbook(m_path, read_only=True); ws = wb.active
                for row in ws.iter_rows(min_row=2, max_row=8000, values_only=True): 
                    if row and len(row) >= 9 and str(row[5]).strip().upper() == part:
                        matches.append({'project': row[0], 'country': row[1], 'batch': row[2], 'assembly': row[3], 'draw': row[4], 'part': row[5], 'rev': row[6], 'total': row[7], 'eng': row[8]})
                wb.close()
            for ui_proj in PROJECTS:
                sc = PROJ_MAP.get(ui_proj); p_f = os.path.join(BASE_PATH, f"{ui_proj}.xlsx")
                if os.path.exists(p_f):
                    wb = load_workbook(p_f, read_only=True)
                    for sn in wb.sheetnames:
                        ws = wb[sn]
                        for row in ws.iter_rows(min_row=3, values_only=True): 
                            if row and len(row) >= 5 and str(row[2]).strip().upper() == part:
                                if not any(m['project'] == sc and str(m['rev']) == str(row[3]) for m in matches):
                                    matches.append({'project': sc, 'country': "Tanzania" if "TRC" in ui_proj else "Malaysia", 'batch': "-", 'assembly': sn, 'draw': row[1], 'part': row[2], 'rev': row[3], 'total': row[4], 'eng': "-"})
                    wb.close()
        except: pass
        
        matches.sort(key=lambda m: (try_int(m['rev'])), reverse=True)
        self.after(0, lambda: self._post_check(matches, part))

    def _post_check(self, matches, part_searched=None):
        self.btn_check_part.configure(state="normal", text="Search")
        self.lbl_feedback.configure(text="", fg_color="transparent") 
        if not matches:
             if part_searched: self.trigger_feedback("info", "Part Not Found")
             return
        unique_results = []
        seen_keys = set()
        for m in matches:
            key = (m['project'], m['draw'])
            if key not in seen_keys: seen_keys.add(key); unique_results.append(m)
        if len(unique_results) == 1: self.show_match_found_dialog(unique_results[0])
        else: self.open_selection_window(unique_results)

    def show_match_found_dialog(self, m):
        top = ctk.CTkToplevel(self); top.title("Record Found"); top.geometry("450x400"); top.transient(self); top.grab_set()
        ctk.CTkLabel(top, text="EXISTING RECORD FOUND", font=("Segoe UI", 16, "bold"), text_color=COLOR_SUCCESS).pack(pady=15)
        info_frame = ctk.CTkFrame(top, fg_color="white", border_width=2, border_color="#BDC3C7", corner_radius=10); info_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(info_frame, text="Project:", font=("Segoe UI", 12), text_color="gray").grid(row=0, column=0, padx=15, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text=m['project'], font=("Segoe UI", 14, "bold"), text_color=COLOR_ACCENT).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text="Assembly:", font=("Segoe UI", 12), text_color="gray").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text=m['assembly'], font=("Segoe UI", 14, "bold"), text_color=COLOR_ACCENT).grid(row=1, column=1, padx=5, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text="Drawing:", font=("Segoe UI", 12), text_color="gray").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text=m['draw'], font=("Segoe UI", 14, "bold"), text_color=COLOR_ACCENT).grid(row=2, column=1, padx=5, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text="Revision:", font=("Segoe UI", 12), text_color="gray").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text=m['rev'], font=("Segoe UI", 14, "bold"), text_color=COLOR_ACCENT).grid(row=3, column=1, padx=5, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text="Part No:", font=("Segoe UI", 12), text_color="gray").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text=m['part'], font=("Segoe UI", 14, "bold"), text_color=COLOR_ACCENT).grid(row=4, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkButton(top, text="YES (Auto-Fill)", fg_color=COLOR_SUCCESS, command=lambda: [self.autofill_form(m), top.destroy()]).pack(pady=10)
        ctk.CTkButton(top, text="CANCEL", fg_color=COLOR_DANGER, command=top.destroy).pack(pady=5)

    def open_selection_window(self, matches):
        top = ctk.CTkToplevel(self); top.title("Search Results"); top.geometry("750x550"); top.transient(self); top.grab_set()
        h = ctk.CTkFrame(top, fg_color=COLOR_PRIMARY, height=70, corner_radius=0); h.pack(fill="x")
        ctk.CTkLabel(h, text="SELECT DATA (Sorted by Latest Revision)", font=("Segoe UI", 18, "bold"), text_color="white").pack(pady=15)
        s = ctk.CTkScrollableFrame(top, fg_color="transparent"); s.pack(fill="both", expand=True, padx=20, pady=10)
        for m in matches:
            c = ctk.CTkFrame(s, fg_color="white", border_width=2, border_color="#BDC3C7", corner_radius=8); c.pack(fill="x", pady=8, padx=5, ipady=5)
            ctk.CTkLabel(c, text=f"PROJECT: {m['project']} ({m.get('country', 'N/A')})", font=("Segoe UI", 13, "bold"), text_color=COLOR_ACCENT).pack(side="left", padx=10)
            ctk.CTkButton(c, text="SELECT", width=90, fg_color=COLOR_SUCCESS, command=lambda data=m: [self.autofill_form(data), top.destroy()]).pack(side="right", padx=10)
        ctk.CTkButton(top, text="CANCEL", fg_color=COLOR_DANGER, width=120, command=top.destroy).pack(pady=15)

    def autofill_form(self, data):
        sc = str(data.get('project')).strip(); ci = str(data.get('country', '')).upper()
        ui = "H10 TRC" if sc == "H10" and "TANZANIA" in ci else "H10 BeraPit" if sc == "H10" else REVERSE_PROJ_MAP.get(sc)
        if ui: self.proj_v.set(ui); self.update_logic(ui) 
        if data.get('batch'): self.batch_v.set("-" if str(data.get('batch')) in [None, "", "None"] else str(data.get('batch')))
        if data.get('assembly'): self.assembly_v.set(str(data.get('assembly')))
        if data.get('draw'): self.draw_ent.delete(0, 'end'); self.draw_ent.insert(0, str(data.get('draw')))
        if data.get('rev') is not None: self.rev_ent.delete(0, 'end'); self.rev_ent.insert(0, str(data.get('rev'))); self.auto_remark_logic(None)
        if data.get('total') is not None: self.total_ent.delete(0, 'end'); self.total_ent.insert(0, str(data.get('total')))
        if data.get('eng') and data.get('eng') in ENGINEER_LIST: self.eng_v.set(data.get('eng'))

    # --- SUBMIT & DUPLICATE LOGIC ---
    def check_duplicate_entry(self, fp, tsn, pi, ri):
        if not os.path.exists(fp): return False 
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
            sn = next((s for s in wb.sheetnames if s.strip().lower() == tsn.lower()), None)
            if not sn: wb.close(); return False 
            ws = wb[sn]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and len(row) >= 4 and str(row[2]).strip().upper() == pi and str(row[3]) == str(ri):
                    wb.close(); return True
            wb.close(); return False
        except: return False 

    def check_master_duplicate(self, cp):
        mp = os.path.join(BASE_PATH, MASTER_FILE)
        if not os.path.exists(mp): return False
        try:
            wb = load_workbook(mp, read_only=True, data_only=True); ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=8000, values_only=True):
                if not row or len(row) < 9: continue
                match = True
                for i in range(9): 
                    if str(row[i]).strip().upper() != str(cp[i]).strip().upper():
                        match = False; break
                if match: wb.close(); return True 
            wb.close(); return False
        except: return False

    def show_duplicate_choice_dialog(self, part, rev, sql_data, excel_master_data, master_file_path):
        top = ctk.CTkToplevel(self); top.title("Action Required"); top.geometry("480x320"); top.transient(self); top.grab_set()
        ctk.CTkLabel(top, text="DUPLICATE IN PROJECT EXCEL", font=FONT_SECTION, text_color=COLOR_DANGER).pack(pady=15)
        ctk.CTkLabel(top, text=f"Part No: {part} (Rev {rev}) already exists in sub-project.").pack(pady=5)
        btn_f = ctk.CTkFrame(top, fg_color="transparent"); btn_f.pack(pady=15)
        def h_master():
            top.destroy()
            try:
                wb = load_workbook(master_file_path); ws = wb.active
                d = list(excel_master_data); d[10] = "Re-Release"; ws.append(d)
                lr = ws.max_row
                for idx in range(1, 12):
                    c = ws.cell(row=lr, column=idx)
                    if idx == 10: c.number_format = 'DD/MM/YYYY'
                    c.alignment = LEFT_ALIGN if idx in [4, 5, 6] else CENTER_ALIGN
                    c.border = THIN_BORDER
                wb.save(master_file_path); self.trigger_feedback("success", "Logged to Master Only")
            except Exception as e: messagebox.showerror("Error", str(e))
        def h_sql():
            top.destroy()
            try:
                conn = psycopg2.connect(**DB_CONFIG); cur = conn.cursor()
                cur.execute("""SELECT 1 FROM project_data WHERE project_name=%s AND country=%s AND batch=%s AND main_assembly=%s AND drawing_name=%s AND part_number=%s AND revision=%s AND total_sheets=%s AND engineer=%s""", 
                            (sql_data[0], sql_data[1], sql_data[2], sql_data[3], sql_data[4], sql_data[5], sql_data[6], sql_data[7], sql_data[8]))
                if not cur.fetchone():
                    cur.execute("INSERT INTO project_data (project_name, country, batch, main_assembly, drawing_name, part_number, revision, total_sheets, engineer, date_approved, remarks) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", sql_data)
                    conn.commit(); self.trigger_feedback("success", "SQL Only Inserted")
                else: self.update_status("Already in SQL", COLOR_WARNING)
                cur.close(); conn.close()
            except Exception as e: messagebox.showerror("DB Error", str(e))
        ctk.CTkButton(btn_f, text="RE-RELEASE\n(MASTER ONLY)", fg_color=COLOR_PRIMARY, width=150, height=50, command=h_master).grid(row=0, column=0, padx=10)
        ctk.CTkButton(btn_f, text="SQL ONLY\n(DATABASE)", fg_color=COLOR_INFO, width=150, height=50, command=h_sql).grid(row=0, column=1, padx=10)
        ctk.CTkButton(top, text="CANCEL", fg_color=COLOR_DANGER, width=80, command=top.destroy).pack(pady=15)

    def submit(self):
        self.err_draw.configure(text="")
        self.err_part.configure(text="")
        self.err_rev.configure(text="")
        self.err_total.configure(text="")
        full_name = self.proj_v.get(); short_proj = PROJ_MAP.get(full_name)
        proj_file = os.path.join(BASE_PATH, f"{full_name}.xlsx"); master_file_path = os.path.join(BASE_PATH, MASTER_FILE)
        draw = self.draw_ent.get().upper().strip(); part = self.part_ent.get().upper().strip(); rev_str = self.rev_ent.get().strip(); total_str = self.total_ent.get().strip()
        has_err = False
        if not draw: self.err_draw.configure(text="* Sila isi Drawing Name"); has_err = True
        if not part: self.err_part.configure(text="* Sila isi Part Number"); has_err = True
        if not rev_str: self.err_rev.configure(text="* Sila isi Revision"); has_err = True
        if not total_str: self.err_total.configure(text="* Sila isi Total Sheets"); has_err = True
        if has_err: self.trigger_feedback("error", "Isi ruangan bertanda merah!"); return
        try: rev, total = int(rev_str), int(total_str)
        except: self.trigger_feedback("error", "Revision/Total must be numbers"); return

        self.set_process_highlight("💾 PROCESSING DATA... PLEASE DO NOT CLOSE", COLOR_WARNING)
        tsn = self.assembly_v.get()[:31]; batch_val = self.batch_v.get(); bve = int(batch_val) if batch_val.isdigit() else batch_val
        dt_obj = self.date_picker.get_date(); dt_sql = dt_obj.strftime('%Y-%m-%d'); country = "Tanzania" if full_name == "H10 TRC" else "Malaysia"

        for f in [proj_file, master_file_path]:
            if os.path.exists(f):
                try: 
                    with open(f, "a"): pass
                except PermissionError: 
                    self.lbl_feedback.configure(text="", fg_color="transparent")
                    messagebox.showerror("Error", f"Close {os.path.basename(f)}!"); return

        sql_data = [short_proj, country, batch_val, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get(), dt_sql, self.remark_v.get()]
        excel_master_data = [short_proj, country, bve, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get(), dt_obj, self.remark_v.get()]

        if self.check_duplicate_entry(proj_file, tsn, part, rev):
            self.lbl_feedback.configure(text="", fg_color="transparent")
            self.show_duplicate_choice_dialog(part, rev, sql_data, excel_master_data, master_file_path); return

        try:
            conn = psycopg2.connect(**DB_CONFIG); cur = conn.cursor()
            cur.execute("""SELECT 1 FROM project_data WHERE project_name=%s AND country=%s AND batch=%s AND main_assembly=%s AND drawing_name=%s AND part_number=%s AND revision=%s AND total_sheets=%s AND engineer=%s""", 
                        (short_proj, country, batch_val, self.assembly_v.get(), draw, part, rev, total, self.eng_v.get()))
            if not cur.fetchone():
                cur.execute("INSERT INTO project_data (project_name, country, batch, main_assembly, drawing_name, part_number, revision, total_sheets, engineer, date_approved, remarks) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", sql_data)
                conn.commit()
            cur.close(); conn.close()

            if not os.path.exists(master_file_path):
                wb_m = Workbook(); ws_m = wb_m.active; ws_m.title = "MasterList"; ws_m.append(MASTER_HEADERS); wb_m.save(master_file_path)
            add_to_master = True; final_master_data = list(excel_master_data)
            if self.check_master_duplicate(excel_master_data):
                self.lbl_feedback.configure(text="", fg_color="transparent")
                if messagebox.askyesno("Master Log", "Matches Master exactly. Re-Release?"): final_master_data[10] = "Re-Release"
                else: add_to_master = False
                self.set_process_highlight("💾 CONTINUING WRITE PROCESS...", COLOR_WARNING)
            if add_to_master:
                wb_m = load_workbook(master_file_path); ws_m = wb_m.active; ws_m.append(final_master_data); lr = ws_m.max_row
                for idx in range(1, 12):
                    cell = ws_m.cell(row=lr, column=idx)
                    if idx == 10: cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = LEFT_ALIGN if idx in [4, 5, 6] else CENTER_ALIGN
                    cell.border = THIN_BORDER
                wb_m.save(master_file_path)

            wb_p = load_workbook(proj_file)
            asn = next((s for s in wb_p.sheetnames if s.strip().lower() == tsn.lower()), None)
            if not asn:
                ws_p = wb_p.create_sheet(tsn); ws_p.append(["Document Control"]); ws_p.append(["Sl. No", "Drawing Name", "Part Number", "Revision", "Total Drawings", "Date Approved", "Remarks"])
                ws_p.cell(row=15, column=1, value="Drawing issued by:-"); dsr = 3
            else:
                ws_p = wb_p[asn]; dsr = 3
                for r in range(1, 11):
                    if "part number" in str(ws_p.cell(row=r, column=3).value or "").lower(): dsr = r + 1; break
            sig = next((r for r in range(dsr, ws_p.max_row + 50) if any("issued by" in str(ws_p.cell(row=r, column=c).value or "").lower() for c in range(1, 6))), None)
            if not sig: sig = max(ws_p.max_row + 2, 4); ws_p.cell(row=sig, column=1, value="Drawing issued by:-")
            tr = next((r for r in range(dsr, sig) if str(ws_p.cell(row=r, column=3).value or "").strip().upper() == part), None)
            is_ov = tr is not None; rts = "Revised" if is_ov else self.remark_v.get()
            if not tr:
                ldr = dsr - 1
                for r in range(max(dsr, sig - 1), dsr - 1, -1):
                    if ws_p.cell(row=r, column=2).value or ws_p.cell(row=r, column=3).value: ldr = r; break
                tr = ldr + 1; ws_p.insert_rows(tr)
                for r in range(ws_p.max_row, tr, -1):
                    if (r-1) in ws_p.row_dimensions: ws_p.row_dimensions[r].height = ws_p.row_dimensions[r-1].height
            sl_no = tr - dsr + 1 if not is_ov else ws_p.cell(row=tr, column=1).value
            fd = [sl_no, draw, part, rev, total, dt_obj, rts]
            for col, val in enumerate(fd, 1):
                cell = ws_p.cell(row=tr, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = val
                    if col == 6: cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = LEFT_ALIGN if col in [2, 3] else CENTER_ALIGN
                    cell.border = THIN_BORDER
            sig_rc = next((r for r in range(dsr, ws_p.max_row + 10) if any("issued by" in str(ws_p.cell(row=r, column=c).value or "").lower() for c in range(1, 6))), ws_p.max_row + 1)
            dm = []
            if sig_rc - 1 >= dsr:
                for r in range(dsr, sig_rc):
                    rd = [ws_p.cell(row=r, column=c).value for c in range(2, 8)]
                    if any(rd): dm.append(rd)
                dm.sort(key=lambda x: str(x[1]).strip().upper() if x[1] else "")
                for i, rd in enumerate(dm):
                    cr = dsr + i; slc = ws_p.cell(row=cr, column=1); slc.value = i+1; slc.alignment = CENTER_ALIGN; slc.border = THIN_BORDER
                    for idx, v in enumerate(rd):
                        c = idx + 2; cell = ws_p.cell(row=cr, column=c); cell.value = v
                        if c == 6: cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = LEFT_ALIGN if c in [2, 3] else CENTER_ALIGN
                        cell.border = THIN_BORDER
            dob = []
            for r in range(dsr, sig_rc):
                dv = ws_p.cell(row=r, column=6).value
                if isinstance(dv, (datetime, date)): dob.append(dv.date() if isinstance(dv, datetime) else dv)
            ldo = max(dob) if dob else None
            for r in range(dsr, sig_rc):
                dv = ws_p.cell(row=r, column=6).value
                cdo = dv.date() if isinstance(dv, datetime) else dv if isinstance(dv, date) else None
                isl = (cdo == ldo and ldo is not None)
                for c in range(1, 8):
                    cell = ws_p.cell(row=r, column=c)
                    if isinstance(cell, MergedCell): continue
                    if c == 7 and not isl: cell.value = None
                    cell.fill = CREAM_YELLOW if isl else NO_FILL
                    cell.font = Font(bold=isl)
                ws_p.row_dimensions[r].height = 18

            wb_p.save(proj_file)
            self.trigger_feedback("success", f"✅ DATA SUCCESSFULLY WRITTEN TO {os.path.basename(proj_file)}!")
            self.update_status(f"Saved: {part}", COLOR_SUCCESS)
        except Exception as e: 
            self.lbl_feedback.configure(text="", fg_color="transparent")
            self.update_status(f"Error: {e}", COLOR_DANGER)

    # --- PDF SEARCH LOGIC ---
    def search_pdf_thread(self):
        part_no = self.pdf_search_ent.get().strip().upper()
        if not part_no: return
        self.btn_search_pdf.configure(state="disabled", text="Searching...")
        threading.Thread(target=self._run_pdf_search_dynamic, args=(part_no,), daemon=True).start()

    def _run_pdf_search_dynamic(self, part_no):
        found_matches = []
        master_records = []
        master_path = os.path.join(BASE_PATH, MASTER_FILE)
        if os.path.exists(master_path):
            try:
                wb_m = load_workbook(master_path, data_only=True, read_only=True)
                ws_m = wb_m.active
                for row in ws_m.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 9 and str(row[5]).strip().upper() == part_no:
                        master_records.append({"project": row[0], "assy": row[3], "draw": row[4], "part": row[5], "rev": row[6], "total": row[7]})
                wb_m.close()
            except: pass
        if not os.path.exists(BASE_DRAWINGS_PATH):
            self.after(0, lambda: self.pdf_status_lbl.configure(text="❌ Drive Y: unreachable.", text_color=COLOR_DANGER))
            self.after(0, lambda: self.btn_search_pdf.configure(state="normal", text="SEARCH DRAWING"))
            return
        try:
            for proj_folder in os.listdir(BASE_DRAWINGS_PATH):
                proj_path = os.path.join(BASE_DRAWINGS_PATH, proj_folder)
                if not os.path.isdir(proj_path): continue
                for assy_folder in os.listdir(proj_path):
                    assy_path = os.path.join(proj_path, assy_folder)
                    if not os.path.isdir(assy_path): continue
                    latest_file = os.path.join(assy_path, f"{part_no}.pdf")
                    if os.path.exists(latest_file):
                        meta = self._find_best_meta(part_no, proj_folder, assy_folder, master_records, is_latest=True)
                        found_matches.append({**meta, "path": latest_file, "type": "Latest"})
                    for item in os.listdir(assy_path):
                        item_path = os.path.join(assy_path, item)
                        if os.path.isdir(item_path) and "obsolete" in item.lower():
                            for obs_file in os.listdir(item_path):
                                if obs_file.upper().startswith(part_no) and obs_file.lower().endswith(".pdf"):
                                    rev_f = self._extract_rev_from_filename(obs_file)
                                    meta = self._find_best_meta(part_no, proj_folder, assy_folder, master_records, rev_to_match=rev_f)
                                    found_matches.append({**meta, "path": os.path.join(item_path, obs_file), "type": "Obsolete"})
        except: pass
        self.after(0, lambda: self._handle_pdf_results(found_matches))

    def _find_best_meta(self, part, proj, assy, records, is_latest=False, rev_to_match=None):
        default = {"part": part, "draw": "Unknown Title", "project": proj, "assy": assy, "rev": "N/A", "total": "N/A"}
        short_proj = PROJ_MAP.get(proj, proj)
        matches = [r for r in records if r["project"] == short_proj]
        if not matches: return default
        if is_latest:
            matches.sort(key=lambda x: try_int(x["rev"]), reverse=True)
            return matches[0]
        elif rev_to_match is not None:
            for r in matches:
                if str(r["rev"]) == str(rev_to_match): return r
        return matches[0]

    def _extract_rev_from_filename(self, filename):
        try:
            if "REV" in filename.upper(): return filename.upper().split("REV")[-1].strip().split(".")[0].strip()
        except: pass
        return None

    def _handle_pdf_results(self, results):
        self.btn_search_pdf.configure(state="normal", text="SEARCH DRAWING")
        if not results:
            self.pdf_status_lbl.configure(text="❌ No record found on server.", text_color=COLOR_DANGER)
            messagebox.showwarning("No Records Found", f"No signed drawings found for {self.pdf_search_ent.get()} on server.")
            return
        self.pdf_status_lbl.configure(text=f"✅ Found {len(results)} drawing(s).", text_color=COLOR_SUCCESS)
        self.show_pdf_selection_dialog(results)

    def show_pdf_selection_dialog(self, results):
        """Dialog untuk memilih drawing PDF dengan butiran lengkap dan highlight LATEST."""
        top = ctk.CTkToplevel(self); top.title("Drawing Records Found"); top.geometry("700x550"); top.transient(self); top.grab_set()
        
        ctk.CTkLabel(top, text="SIGNED DRAWINGS FOUND", font=("Segoe UI", 18, "bold"), text_color=COLOR_ACCENT).pack(pady=15)
        ctk.CTkLabel(top, text="Select a record to open the Signed Copy PDF:", font=("Segoe UI", 12)).pack(pady=5)

        scroll = ctk.CTkScrollableFrame(top, width=640, height=350, fg_color="transparent")
        scroll.pack(padx=20, pady=10, fill="both", expand=True)

        for res in results:
            is_latest = res["type"] == "Latest"
            # Highlight border if Latest
            card_border_color = COLOR_SUCCESS if is_latest else "#BDC3C7"
            card_border_width = 3 if is_latest else 1
            
            card = ctk.CTkFrame(scroll, fg_color="white", border_width=card_border_width, border_color=card_border_color, corner_radius=10)
            card.pack(fill="x", pady=8, padx=5)
            
            # Info Layout
            main_f = ctk.CTkFrame(card, fg_color="transparent")
            main_f.pack(side="left", padx=15, pady=12, fill="both", expand=True)
            
            # Badge LATEST if applicable
            if is_latest:
                badge_f = ctk.CTkFrame(main_f, fg_color=COLOR_SUCCESS, corner_radius=5)
                badge_f.pack(anchor="w", pady=(0, 5))
                ctk.CTkLabel(badge_f, text=" LATEST REVISION ", font=("Segoe UI", 10, "bold"), text_color="white").pack(padx=5, pady=2)

            # Row 1: Part Number & Drawing Name
            row1 = ctk.CTkFrame(main_f, fg_color="transparent")
            row1.pack(fill="x")
            ctk.CTkLabel(row1, text=res["part"], font=("Segoe UI", 14, "bold"), text_color=COLOR_PRIMARY).pack(side="left")
            ctk.CTkLabel(row1, text=f"  |  {res['draw']}", font=("Segoe UI", 13), text_color=COLOR_TEXT).pack(side="left")

            # Row 2: Project & Assembly
            row2 = ctk.CTkFrame(main_f, fg_color="transparent")
            row2.pack(fill="x", pady=(2, 0))
            ctk.CTkLabel(row2, text=f"Project: {res['project']}   •   Assembly: {res['assy']}", font=("Segoe UI", 11), text_color="gray").pack(side="left")

            # Row 3: Revision & Total Sheets
            row3 = ctk.CTkFrame(main_f, fg_color="transparent")
            row3.pack(fill="x")
            ctk.CTkLabel(row3, text=f"Revision: {res['rev']}   •   Sheets: {res['total']}", font=("Segoe UI", 11, "bold"), text_color=COLOR_ACCENT).pack(side="left")

            # Open Button
            btn_color = COLOR_SUCCESS if is_latest else "#95a5a6"
            ctk.CTkButton(card, text="OPEN PDF", width=110, height=40, fg_color=btn_color, font=("Segoe UI", 12, "bold"), 
                          command=lambda p=res["path"]: os.startfile(p)).pack(side="right", padx=15)

        ctk.CTkButton(top, text="CLOSE WINDOW", fg_color=COLOR_DANGER, width=150, height=40, font=("Segoe UI", 12, "bold"), command=top.destroy).pack(pady=20)

    def publish_to_server_thread(self):
        self.btn_publish.configure(state="disabled", text="Publishing...")
        threading.Thread(target=self._run_publish_logic, daemon=True).start()

    def _run_publish_logic(self):
        if not os.path.exists(SERVER_PATH):
            self.after(0, lambda: messagebox.showerror("Error", "Server unreachable")); return
        suc, fail = 0, 0
        for l, s in FILE_MAPPING.items():
            lf, sf = os.path.join(BASE_PATH, l), os.path.join(SERVER_PATH, s)
            if not os.path.exists(lf): fail += 1; continue
            try:
                if os.path.exists(sf): os.rename(sf, sf)
                shutil.copy2(lf, sf); suc += 1
            except: fail += 1
        self.after(0, self._finalize_publish, suc, fail)

    def _finalize_publish(self, s, f):
        self.lbl_last_sync.configure(text=f"Last Sync: {datetime.now().strftime('%H:%M:%S')}")
        self.btn_publish.configure(state="normal", text="PUBLISH TO SERVER")
        messagebox.showinfo("Result", f"Success: {s} | Fail: {f}")

def try_int(val):
    try: return int(val)
    except: return -1

if __name__ == "__main__":
    app = DCDEApp(); app.mainloop()